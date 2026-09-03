"""產生最終報告 Excel：BBG快照 / 遠期走勢圖 / 原始數據 / 合併版面。

圖表引擎由 ``chart.engine`` 切換：
- ``matplotlib``：靜態 PNG 嵌入 openpyxl（預設）
- ``xlsxwriter``：Excel 原生可互動折線圖

「遠期走勢圖」與「原始數據」的資料來源一律是 ``vba_dir / yyyymmdd.xlsx``，
不跟隨 ``output_dir`` / ``run_dir``。原始數據 sheet 貼入靜態值，不使用外部連結。
"""

from __future__ import annotations

import logging
import shutil
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.units import (
    DEFAULT_ROW_HEIGHT,
    EMU_to_pixels,
    pixels_to_EMU,
    points_to_pixels,
)
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet

from lme_daily.bbg_fetch import normalize_bbg_cell, value_from_stored_number
from lme_daily.config import AppConfig
from lme_daily.exceptions import ReportBuildError

logger = logging.getLogger(__name__)

DATE_COL = "Prompt"
SIX_SERIES = {
    "CA": "銅 Copper",
    "AH": "鋁 Aluminium",
    "ZS": "鋅 Zinc",
    "NI": "鎳 Nickel",
    "SN": "錫 Tin",
    "PB": "鉛 Lead",
}

SHEET_BBG = "BBG快照"
SHEET_CHART = "遠期走勢圖"
SHEET_RAW = "原始數據"
SHEET_PRINT = "合併版面"
SHEET_CHART_DATA = "_chart_data"
PRINT_RAW_SECTION_TITLE = "Prompt date settlement price"

# BBG快照 / 遠期走勢圖 / 原始數據 sheet 仍用深藍金，避免動到非合併版面頁。
COLOR_NAVY = "0B2447"
COLOR_GOLD = "C5A059"
COLOR_GRAY = "F2F4F7"
COLOR_LINE = "D6D9DE"
COLOR_MUTED = "6B7280"
COLOR_INK = "1F2937"
COLOR_WHITE = "FFFFFF"

# 合併版面樣式（量測自 Fu-Ben-LMEMei-Ri-Bao-Jia-20260827-2.xlsx，勿另行發明數值）
COLOR_ACCENT = "B6895F"  # 主色：標題列/區塊標題列/表頭列底色
COLOR_TEXT_CREAM = "FFF5E8"  # 深色底上的標題文字色
COLOR_TEXT_WHITE = "FFFFFF"  # 表頭代碼列（C8:H8）文字色
COLOR_DATE_BG = "F2F4F7"  # 日期資訊列底色
COLOR_DATE_TEXT = "2B2B2B"  # 日期資訊列文字色
COLOR_DISCLAIMER_TEXT = "1F2937"  # 免責聲明文字色
FONT_NAME = "宋体"
# Excel 預設字型（Calibri 11 / 96 DPI）欄寬字元 → 像素：MDW=7、左右邊界+格線=5
EXCEL_MAX_DIGIT_WIDTH_PX = 7
EXCEL_COLUMN_PADDING_PX = 5
EXCEL_DEFAULT_COLUMN_WIDTH = 8.43
PRINT_LAST_COL = 8  # A:H
PRINT_LOGO_MERGE_RANGE = "A2:H2"
PRINT_CHART_ROW_STRIDE = 13
PRINT_CHART_WIDTH_EMU = 3683635
PRINT_CHART_HEIGHT_EMU = 2402205
EMU_PER_PIXEL = 9525
PRINT_LOGO_ROW_HEIGHT = 34
PRINT_LOGO_ANCHOR = "A2"
# Excel 列高為 point；openpyxl Image 寬高為 pixel（96 DPI）
PRINT_LOGO_PX_PER_POINT = 96 / 72
PRINT_TITLE_ROW_HEIGHT = 34
PRINT_SUBTITLE_ROW_HEIGHT = 16
PRINT_DATE_ROW_HEIGHT = 18
PRINT_SPACER_ROW_HEIGHT = 1
PRINT_SECTION_ROW_HEIGHT = 18

PRINT_FOOTER = "第 &P 頁，共 &N 頁"
PRINT_DISCLAIMER_FULL = (
    "免責聲明：本報告由山證國際金融控股有限公司（\"本公司\"或\"山證國際\"）提供，"
    "所載之內容或意見乃根據本公司認為可靠之數據源來編制，惟本公司並不就此等內容之準確性、"
    "完整性及正確性作出明示或默示之保證。本報告的作用純粹為提供信息，並不應視為對本報告內"
    "提及的任何產品買賣或交易之專業推介、建議、邀請或要約。"
    "\n\n"
    "投資附帶風險，投資者需注意投資項目之價值可升亦可跌，而過往之表現亦不一定反映未來之表現。"
    "投資者進行投資前請尋求獨立之投資意見。本公司竭力確保其提供之數據準確可靠，但不保證該等"
    "數據絕對正確可靠；對於任何因資料不確或遺漏又或因根據或倚賴本資料所作決定、行動或不行動"
    "而引致之損失或損害，本公司概不負責（不論是民事侵權行為責任或合約責任或其他）。"
)
DISCLAIMER_FONT_SIZE = 8
DISCLAIMER_ROW_HEIGHT = 78
CHART_ROW_STRIDE = 19  # 獨立「遠期走勢圖」sheet；合併版面用 PRINT_CHART_ROW_STRIDE
PAPERSIZE_A3 = 8
PAPERSIZE_A4 = 9
# Excel 頁首/頁尾 ``&nn`` 必須兩位數字。``&8`` 接 ``2026-08-20`` 會變成 ``&82``（82pt）。
HF_FONT_SIZE = "08"

EXCEL_EPOCH = datetime(1899, 12, 30)
SOURCE_LABEL = "資料來源"

COMPANY_ZH = "山證國際金融控股有限公司"
COMPANY_EN = "Shanxi Securities International"
REPORT_TITLE_ZH = "LME每日報價"
REPORT_TITLE_EN = "LME Daily Quotation"
NUMBER_FORMAT_2DP = "#,##0.00"
AUTOFIT_PADDING = 3
# 下限：容納 5 位數 + 小數點 + 2 位小數（如 16554.08 / 16,554.08）；不是上限。
AUTOFIT_MIN_WIDTH = 10

PRINT_BBG_METAL_CODES = ("CA", "AH", "PB", "NI", "SN", "ZS")


class MergedLayoutStyle:
    """合併版面唯一樣式來源。數值量測自手動美化檔，所有 print-sheet 寫入都從這裡讀。"""

    accent = COLOR_ACCENT
    text_cream = COLOR_TEXT_CREAM
    text_white = COLOR_TEXT_WHITE
    date_bg = COLOR_DATE_BG
    date_text = COLOR_DATE_TEXT
    disclaimer_text = COLOR_DISCLAIMER_TEXT
    font_name = FONT_NAME
    last_col = PRINT_LAST_COL
    chart_row_stride = PRINT_CHART_ROW_STRIDE
    chart_width_emu = PRINT_CHART_WIDTH_EMU
    chart_height_emu = PRINT_CHART_HEIGHT_EMU
    disclaimer_font_size = DISCLAIMER_FONT_SIZE
    disclaimer_row_height = DISCLAIMER_ROW_HEIGHT
    row_logo_height = PRINT_LOGO_ROW_HEIGHT
    row_title_height = PRINT_TITLE_ROW_HEIGHT
    row_subtitle_height = PRINT_SUBTITLE_ROW_HEIGHT
    row_date_height = PRINT_DATE_ROW_HEIGHT
    row_spacer_height = PRINT_SPACER_ROW_HEIGHT
    row_section_height = PRINT_SECTION_ROW_HEIGHT
    size_company = 11
    size_title = 20
    size_subtitle = 11
    size_date = 11
    size_section = 12
    size_code_header = 11
    size_body = 11
    size_disclaimer = DISCLAIMER_FONT_SIZE

    @classmethod
    def font(cls, *, size: float, bold: bool = False, color: str | None = None) -> Font:
        kwargs: dict[str, Any] = {"name": cls.font_name, "size": size, "bold": bold}
        if color:
            kwargs["color"] = color
        return Font(**kwargs)

    @classmethod
    def fill(cls, color: str) -> PatternFill:
        return PatternFill("solid", fgColor=color)

    @classmethod
    def thin_border(cls) -> Border:
        side = Side(style="thin", color=COLOR_LINE)
        return Border(left=side, right=side, top=side, bottom=side)

    @classmethod
    def bottom_border(cls) -> Border:
        side = Side(style="thin", color=COLOR_LINE)
        return Border(bottom=side)

    @classmethod
    def center(cls) -> Alignment:
        return Alignment(horizontal="center", vertical="center")

    @classmethod
    def last_col_letter(cls) -> str:
        return get_column_letter(cls.last_col)

    @classmethod
    def chart_width_px(cls) -> int:
        return round(cls.chart_width_emu / EMU_PER_PIXEL)

    @classmethod
    def chart_height_px(cls) -> int:
        return round(cls.chart_height_emu / EMU_PER_PIXEL)

    @classmethod
    def chart_anchors(cls, start_row: int) -> list[str]:
        anchors: list[str] = []
        for pair in range(3):
            row = start_row + pair * cls.chart_row_stride
            anchors.extend((f"A{row}", f"D{row}"))
        return anchors

    @classmethod
    def xw_format(cls, workbook: Any, **extra: Any) -> Any:
        payload = {"font_name": cls.font_name, **extra}
        return workbook.add_format(payload)


def _css(color: str) -> str:
    return color if color.startswith("#") else f"#{color}"


def _win_path(path: Path) -> str:
    return str(path).replace("/", "\\")


def header_footer_run(color: str, text: str, *, size: str = HF_FONT_SIZE) -> str:
    """Excel 頁首/頁尾片段：顏色 + 兩位字級 + 文字。

    ``&nn`` 後面加空白，避免日期開頭數字被吃進字級碼（``&8`` + ``2026`` → ``&82``）。
    """
    return f"&K{color}&{size} {text}"


def _is_excel_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def excel_absolute_external_formula(source_path: Path, sheet_name: str, cell_addr: str) -> str:
    """絕對路徑外部連結：``='C:\\folder\\[file.xlsx]Sheet'!A1``。

    報告若寫在 ``run_dir``（與 VBA 中繼檔不同資料夾），相對連結會斷；一律用絕對路徑。
    """
    folder = _win_path(source_path.parent).rstrip("\\")
    filename = source_path.name
    escaped = sheet_name.replace("'", "''")
    return f"='{folder}\\[{filename}]{escaped}'!{cell_addr}"


def build_report(
    config: AppConfig,
    *,
    as_of: date,
    step2_path: Path,
    bbg_values: tuple[tuple[Any, ...], ...],
    bbg_formats: tuple[tuple[str, ...], ...],
    output_path: Path | None = None,
) -> Path:
    """依 config.chart.engine 產出 ``LME每日報價{yyyymmdd}.xlsx`` 到 ``run_dir``。"""
    dest = output_path or config.output_workbook_path(as_of)
    dest.parent.mkdir(parents=True, exist_ok=True)
    logger.info("開始產生報告（engine=%s）：%s", config.chart.engine, dest)
    logger.info("圖表/原始數據來源（vba_dir 中繼檔）：%s", step2_path)
    logger.debug("%s：%s", SOURCE_LABEL, _win_path(step2_path))

    try:
        step2_ready = step2_path.is_file()
    except OSError as exc:
        raise ReportBuildError(f"無法確認步驟二產出檔：{step2_path}（{exc}）") from exc
    if not step2_ready:
        raise ReportBuildError(
            f"步驟二產出檔不存在：{step2_path}。"
            "「遠期走勢圖」與「原始數據」只讀 vba_dir 下的 yyyymmdd.xlsx，中斷而不產生缺資料報告。"
        )

    curve_df = load_forward_curve(step2_path)
    plot_df = slice_plot_window(curve_df, forward_months=config.chart.forward_months)

    if config.chart.engine == "xlsxwriter":
        _build_with_xlsxwriter(
            dest, step2_path, curve_df, plot_df, bbg_values, bbg_formats, config, as_of
        )
    else:
        _build_with_openpyxl(dest, step2_path, plot_df, bbg_values, bbg_formats, config, as_of)

    if not dest.is_file():
        raise ReportBuildError(f"報告寫入後檔案不存在：{dest}")
    logger.info("報告已寫入：%s", dest)
    return dest


def load_forward_curve(step2_path: Path) -> pd.DataFrame:
    """讀取步驟二單表，統一 Prompt 日期；轉換失敗列記警告但不中斷。"""
    try:
        df = pd.read_excel(step2_path)
    except Exception as exc:
        raise ReportBuildError(f"無法讀取 {step2_path}：{exc}") from exc

    if DATE_COL not in df.columns:
        raise ReportBuildError(
            f"{step2_path.name} 缺少欄位 {DATE_COL!r}。實際欄位：{list(df.columns)}"
        )
    missing_series = [code for code in SIX_SERIES if code not in df.columns]
    if missing_series:
        raise ReportBuildError(
            f"{step2_path.name} 缺少品種欄位 {missing_series}。實際欄位：{list(df.columns)}"
        )

    parsed = pd.to_datetime(df[DATE_COL], dayfirst=True, errors="coerce", format="mixed")
    original_na = df[DATE_COL].isna().sum()
    failed = int(parsed.isna().sum() - original_na)
    if failed:
        bad_rows = df.loc[parsed.isna() & df[DATE_COL].notna(), DATE_COL]
        sample = bad_rows.astype(str).head(8).tolist()
        logger.warning(
            "Prompt 日期轉換失敗 %d 列（不中斷，這些列不會進入圖表）。樣本：%s",
            failed,
            sample,
        )
    out = df.copy()
    out[DATE_COL] = parsed
    if out[DATE_COL].notna().sum() == 0:
        raise ReportBuildError(f"{step2_path.name} 的 {DATE_COL} 全部無法解析為日期")
    logger.info("遠期曲線列數=%d，有效日期=%d", len(out), int(out[DATE_COL].notna().sum()))
    return out


def slice_plot_window(df: pd.DataFrame, *, forward_months: int) -> pd.DataFrame:
    """cash_date = min(Prompt)；只保留 cash_date 起 ``forward_months`` 個月內的列。"""
    valid = df.dropna(subset=[DATE_COL])
    cash_date = valid[DATE_COL].min()
    cutoff = cash_date + relativedelta(months=forward_months)
    plot_df = valid[valid[DATE_COL] <= cutoff].sort_values(DATE_COL)
    logger.info(
        "圖表視窗：cash_date=%s cutoff=%s（+%dmo）列數 %d → %d",
        pd.Timestamp(cash_date).date(),
        pd.Timestamp(cutoff).date(),
        forward_months,
        len(df),
        len(plot_df),
    )
    if plot_df.empty:
        raise ReportBuildError("過濾 27 個月視窗後沒有可畫圖的資料")
    return plot_df.reset_index(drop=True)


def _excel_serial_to_datetime(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value
    if isinstance(value, (int, float)) and 20000 < float(value) < 80000:
        return EXCEL_EPOCH + timedelta(days=float(value))
    return value


def _looks_like_date_format(fmt: str | None) -> bool:
    if not fmt:
        return False
    lowered = fmt.lower()
    return any(token in lowered for token in ("y", "d", "m", "年", "月", "日")) and "h" not in lowered[:3]


def _thin_border() -> Border:
    side = Side(style="thin", color=COLOR_LINE)
    return Border(left=side, right=side, top=side, bottom=side)


def _is_number_like(value: Any) -> bool:
    if value is None or isinstance(value, (bool, str, datetime, date)):
        return False
    if isinstance(value, (int, float)):
        return True
    try:
        float(value)
    except (TypeError, ValueError):
        return False
    return True


def _prepare_written_value(raw: Any, fmt: str | None = None, *, is_header: bool = False) -> tuple[Any, str | None]:
    """normalize 後寫入；數值已在讀取時依 Excel 顯示文字解析。回傳 (value, number_format)。"""
    value = normalize_bbg_cell(raw)
    if is_header:
        return value, None
    if _looks_like_date_format(fmt):
        value = _excel_serial_to_datetime(value)
        return value, "YYYY-MM-DD"
    if _is_number_like(value):
        return value, NUMBER_FORMAT_2DP
    return value, None


def _cell_display_text(value: Any, *, number_format: str | None = None) -> str:
    if value is None:
        return ""
    fmt = number_format or ""
    if _is_number_like(value) and ("0.00" in fmt or fmt in {"", "General", "G"}):
        return f"{float(value):,.2f}"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _visual_len(text: str) -> int:
    return sum(2 if ord(ch) > 127 else 1 for ch in text)


def _autofit_width(max_len: int) -> float:
    if max_len <= 0:
        return 8
    return max(max_len + AUTOFIT_PADDING, AUTOFIT_MIN_WIDTH)


def _merged_positions(ws: Worksheet) -> set[tuple[int, int]]:
    positions: set[tuple[int, int]] = set()
    for rng in ws.merged_cells.ranges:
        positions.update(rng.cells)
    return positions


def apply_autofit_columns(
    ws: Worksheet,
    *,
    padding: int = AUTOFIT_PADDING,
    skip_empty: bool = False,
    skip_merged: bool = True,
    min_col: int | None = None,
    max_col: int | None = None,
) -> None:
    """依內容自動調整欄寬。必須在該 sheet 資料全部寫入後呼叫。

    合併儲存格（頁首／區塊標題／免責聲明）不列入寬度計算，避免把橫向標題
    字串灌進每一欄。``skip_empty`` 時空白分隔欄不改寬度。
    ``min_col`` / ``max_col`` 為 1-based，用來限定例如合併版面 A:H。
    """
    merged = _merged_positions(ws) if skip_merged else set()
    iter_kw: dict[str, int] = {}
    if min_col is not None:
        iter_kw["min_col"] = min_col
    if max_col is not None:
        iter_kw["max_col"] = max_col
    columns = ws.iter_cols(**iter_kw) if iter_kw else ws.columns
    for col_cells in columns:
        max_len = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            if (cell.row, cell.column) in merged:
                continue
            text = _cell_display_text(cell.value, number_format=cell.number_format)
            max_len = max(max_len, _visual_len(text))
        letter = get_column_letter(col_cells[0].column)
        if max_len <= 0:
            if skip_empty:
                continue
            ws.column_dimensions[letter].width = 8
            continue
        fitted = max(max_len + padding, AUTOFIT_MIN_WIDTH)
        current = ws.column_dimensions[letter].width
        if current:
            fitted = max(fitted, float(current))
        ws.column_dimensions[letter].width = fitted


def apply_center_alignment(ws: Worksheet, *, skip_merged: bool = True) -> None:
    """水平＋垂直置中。必須在該 sheet 資料全部寫入後呼叫，只改 alignment。

    合併儲存格（頁首／紅底區塊標題／免責聲明）維持寫入時的對齊。
    """
    center = Alignment(horizontal="center", vertical="center")
    merged = _merged_positions(ws) if skip_merged else set()
    for row in ws.iter_rows():
        for cell in row:
            if (cell.row, cell.column) in merged:
                continue
            cell.alignment = center


def apply_print_data_styles(ws: Worksheet) -> None:
    """合併版面最後一步：數據區塊置中；A:H 一律自動欄寬（含 C/F/G/H 金屬欄）。"""
    apply_center_alignment(ws)
    apply_autofit_columns(
        ws, skip_empty=True, min_col=1, max_col=PRINT_LAST_COL
    )


def column_width_to_pixels(width_chars: float) -> int:
    """Excel 欄寬（字元）→ 像素。MDW=7、padding=5，再用 ``pixels_to_EMU``。"""
    return int(float(width_chars) * EXCEL_MAX_DIGIT_WIDTH_PX + EXCEL_COLUMN_PADDING_PX)


def column_width_to_emu(width_chars: float) -> int:
    return pixels_to_EMU(column_width_to_pixels(width_chars))


def row_height_to_emu(height_points: float) -> int:
    return pixels_to_EMU(points_to_pixels(float(height_points)))


def worksheet_column_width_chars(ws: Worksheet, col_idx: int) -> float:
    letter = get_column_letter(col_idx)
    width = ws.column_dimensions[letter].width
    if width:
        return float(width)
    default = getattr(ws.sheet_format, "defaultColWidth", None)
    if default:
        return float(default)
    return EXCEL_DEFAULT_COLUMN_WIDTH


def worksheet_row_height_points(ws: Worksheet, row_idx: int) -> float:
    height = ws.row_dimensions[row_idx].height
    if height:
        return float(height)
    default = getattr(ws.sheet_format, "defaultRowHeight", None)
    if default:
        return float(default)
    return float(DEFAULT_ROW_HEIGHT)


def _offset_within_span(sizes_emu: list[int], offset_emu: int) -> tuple[int, int]:
    """把總偏移量拆成（從第幾格開始, 該格內的 colOff/rowOff）。避免全部塞進第一格溢位。"""
    remaining = max(0, int(offset_emu))
    if not sizes_emu:
        return 0, remaining
    for index, size in enumerate(sizes_emu):
        size = max(0, int(size))
        if index == len(sizes_emu) - 1 or remaining < size:
            return index, remaining
        remaining -= size
    return len(sizes_emu) - 1, remaining


def compute_center_anchor_in_range(
    *,
    col_widths_chars: list[float],
    row_heights_pt: list[float],
    image_width_px: int,
    image_height_px: int,
    start_col_0: int,
    start_row_0: int,
) -> tuple[int, int, int, int]:
    """回傳置中後的 ``(col_0, colOff_EMU, row_0, rowOff_EMU)``。"""
    col_emus = [column_width_to_emu(width) for width in col_widths_chars]
    row_emus = [row_height_to_emu(height) for height in row_heights_pt]
    total_w = sum(col_emus)
    total_h = sum(row_emus)
    img_w = pixels_to_EMU(max(0, int(image_width_px)))
    img_h = pixels_to_EMU(max(0, int(image_height_px)))
    x_off = max(0, (total_w - img_w) // 2)
    y_off = max(0, (total_h - img_h) // 2)
    col_i, col_off = _offset_within_span(col_emus, x_off)
    row_i, row_off = _offset_within_span(row_emus, y_off)
    return start_col_0 + col_i, col_off, start_row_0 + row_i, row_off


def center_image_in_merged_range(ws: Worksheet, image: XLImage, merge_range: str) -> None:
    """把已設好寬高的浮動圖片置中於合併範圍（OneCellAnchor + EMU 偏移）。"""
    min_col, min_row, max_col, max_row = range_boundaries(merge_range)
    col_widths = [worksheet_column_width_chars(ws, col) for col in range(min_col, max_col + 1)]
    row_heights = [worksheet_row_height_points(ws, row) for row in range(min_row, max_row + 1)]
    img_w = max(1, int(image.width or 1))
    img_h = max(1, int(image.height or 1))
    col_0, col_off, row_0, row_off = compute_center_anchor_in_range(
        col_widths_chars=col_widths,
        row_heights_pt=row_heights,
        image_width_px=img_w,
        image_height_px=img_h,
        start_col_0=min_col - 1,
        start_row_0=min_row - 1,
    )
    marker = AnchorMarker(col=col_0, colOff=int(col_off), row=row_0, rowOff=int(row_off))
    image.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(cx=pixels_to_EMU(img_w), cy=pixels_to_EMU(img_h)),
    )
    ws.add_image(image)


def disclaimer_row_after(last_content_row: int) -> int:
    """最後一筆內容列之後空一行再放免責聲明（1-based）。"""
    return int(last_content_row) + 2


class _ColWidthAcc:
    """xlsxwriter 無法回讀儲存格，寫入時累計各欄顯示寬度。"""

    def __init__(self) -> None:
        self._max: dict[int, int] = {}

    def add(self, col: int, value: Any, number_format: str | None = None) -> None:
        text = _cell_display_text(value, number_format=number_format)
        if not text:
            return
        self._max[col] = max(self._max.get(col, 0), _visual_len(text))

    def fitted_width(self, col: int) -> float | None:
        max_len = self._max.get(col, 0)
        if max_len <= 0:
            return None
        return _autofit_width(max_len)

    def apply_xlsxwriter(self, ws: Any, *, min_col: int | None = None, max_col: int | None = None) -> None:
        cols = sorted(self._max)
        if min_col is not None or max_col is not None:
            lo = min_col if min_col is not None else 0
            hi = max_col if max_col is not None else max(cols, default=-1)
            cols = [col for col in range(lo, hi + 1) if col in self._max]
        for col in cols:
            width = self.fitted_width(col)
            if width is None:
                continue
            ws.set_column(col, col, width)


def _write_bbg_sheet_openpyxl(
    ws: Worksheet,
    values: tuple[tuple[Any, ...], ...],
    formats: tuple[tuple[str, ...], ...],
    *,
    start_row: int = 1,
    header_fill_color: str = COLOR_NAVY,
) -> int:
    ws.sheet_view.showGridLines = True
    header_fill = PatternFill("solid", fgColor=header_fill_color)
    header_font = Font(color=COLOR_WHITE, bold=True)
    last_row = start_row - 1
    for r_idx, row in enumerate(values):
        fmt_row = formats[r_idx] if r_idx < len(formats) else ()
        excel_row = start_row + r_idx
        last_row = excel_row
        for c_idx, raw in enumerate(row):
            fmt = fmt_row[c_idx] if c_idx < len(fmt_row) else None
            value, num_fmt = _prepare_written_value(raw, fmt, is_header=(r_idx == 0))
            cell = ws.cell(row=excel_row, column=c_idx + 1, value=value)
            cell.border = _thin_border()
            if num_fmt:
                cell.number_format = num_fmt
            elif isinstance(value, datetime):
                cell.number_format = "YYYY-MM-DD"
            elif isinstance(value, date):
                cell.number_format = "YYYY-MM-DD"
            if r_idx == 0:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
    return last_row


def _copy_raw_sheet(
    source_path: Path,
    dest_ws: Worksheet,
    *,
    start_row: int = 1,
) -> int:
    """把步驟二中繼檔的「值」貼進目標表，不寫公式、不建外部連結。"""
    src_wb = load_workbook(source_path, data_only=True)
    last_row = start_row - 1
    try:
        src_ws = src_wb.active
        for row in src_ws.iter_rows():
            for cell in row:
                value = cell.value
                if _is_excel_formula(value):
                    logger.debug("原始數據略過公式儲存格 %s", cell.coordinate)
                    continue
                if value is None:
                    continue
                target_row = cell.row + start_row - 1
                last_row = max(last_row, target_row)
                is_header = cell.row == 1
                if not is_header:
                    value = value_from_stored_number(value)
                target = dest_ws.cell(row=target_row, column=cell.column, value=value)
                if _is_number_like(value):
                    target.number_format = NUMBER_FORMAT_2DP
                elif cell.number_format:
                    target.number_format = cell.number_format
                if cell.has_style and cell.font and cell.font.bold:
                    target.font = Font(bold=True)
    finally:
        src_wb.close()
    return last_row


def _apply_print_layout_openpyxl(
    ws: Worksheet,
    *,
    header: str,
    tab_color: str | None = None,
    paper_size: int = PAPERSIZE_A4,
    header_color: str = COLOR_NAVY,
) -> None:
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = paper_size
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.horizontalCentered = True
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddHeader.center.text = header_footer_run(header_color, header)
    ws.oddFooter.center.text = PRINT_FOOTER
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.75, bottom=0.7, header=0.3, footer=0.4
    )
    ws.print_options.horizontalCentered = True
    if tab_color:
        ws.sheet_properties.tabColor = tab_color


def _set_print_area(ws: Worksheet, last_row: int, last_col: int = PRINT_LAST_COL) -> None:
    if last_row < 1:
        return
    ws.print_area = f"A1:{get_column_letter(last_col)}{last_row}"


def _print_fill_range(ws: Worksheet, row: int, last_col: int, fill: PatternFill) -> None:
    for col in range(1, last_col + 1):
        ws.cell(row=row, column=col).fill = fill


def _add_anchored_image(
    ws: Worksheet,
    image_path: Path,
    *,
    anchor: str,
    width: int | None = None,
    height: int | None = None,
) -> None:
    """浮動圖片：``XLImage`` + 可選寬高 + ``add_image(anchor)``。走勢圖與 Logo 共用。"""
    img = XLImage(str(image_path))
    if width is not None:
        img.width = width
    if height is not None:
        img.height = height
    ws.add_image(img, anchor)


def _logo_pixel_size(image_path: Path, *, row_height_pt: float) -> tuple[int, int]:
    """高度對齊列高（point → pixel），寬度依原圖比例縮放，避免變形。"""
    probe = XLImage(str(image_path))
    native_w = max(1, int(probe.width or 1))
    native_h = max(1, int(probe.height or 1))
    target_h = max(1, round(row_height_pt * PRINT_LOGO_PX_PER_POINT))
    target_w = max(1, round(native_w * target_h / native_h))
    return target_w, target_h


def _place_print_logo(
    ws: Worksheet,
    logo_path: Path | None,
    *,
    merge_range: str = PRINT_LOGO_MERGE_RANGE,
) -> None:
    """合併版面 A2:H2 浮動 Logo，置中於合併範圍。路徑未設或檔案不存在時只記 WARNING。"""
    if logo_path is None:
        logger.warning("未設定 Logo 路徑，略過插入")
        return
    try:
        exists = logo_path.is_file()
    except OSError:
        exists = False
    if not exists:
        logger.warning("找不到 Logo 檔案，略過插入：%s", logo_path)
        return
    try:
        width, height = _logo_pixel_size(logo_path, row_height_pt=PRINT_LOGO_ROW_HEIGHT)
        img = XLImage(str(logo_path))
        img.width = width
        img.height = height
        center_image_in_merged_range(ws, img, merge_range)
        logger.info("已插入合併版面 Logo：%s（%dx%d，置中 %s）", logo_path, width, height, merge_range)
    except Exception as exc:
        logger.warning("無法插入 Logo（%s），略過插入：%s", logo_path, exc)


def _place_print_logo_xlsxwriter(
    ws: Any,
    logo_path: Path | None,
    *,
    col_widths_chars: list[float],
    row_height_pt: float = PRINT_LOGO_ROW_HEIGHT,
    start_col_0: int = 0,
    start_row_0: int = 1,
) -> None:
    if logo_path is None:
        logger.warning("未設定 Logo 路徑，略過插入")
        return
    try:
        exists = logo_path.is_file()
    except OSError:
        exists = False
    if not exists:
        logger.warning("找不到 Logo 檔案，略過插入：%s", logo_path)
        return
    try:
        probe = XLImage(str(logo_path))
        native_w = max(1, int(probe.width or 1))
        native_h = max(1, int(probe.height or 1))
        width, height = _logo_pixel_size(logo_path, row_height_pt=row_height_pt)
        col_0, col_off, row_0, row_off = compute_center_anchor_in_range(
            col_widths_chars=col_widths_chars,
            row_heights_pt=[row_height_pt],
            image_width_px=width,
            image_height_px=height,
            start_col_0=start_col_0,
            start_row_0=start_row_0,
        )
        ws.insert_image(
            row_0,
            col_0,
            str(logo_path),
            {
                "x_scale": width / native_w,
                "y_scale": height / native_h,
                "x_offset": EMU_to_pixels(col_off),
                "y_offset": EMU_to_pixels(row_off),
            },
        )
        logger.info("已插入合併版面 Logo：%s", logo_path)
    except Exception as exc:
        logger.warning("無法插入 Logo（%s），略過插入：%s", logo_path, exc)


def _write_print_disclaimer_openpyxl(
    ws: Worksheet, *, last_content_row: int, last_col: int = PRINT_LAST_COL
) -> int:
    """合併版面最下方：完整兩段免責聲明，自動換行、深灰小字。"""
    style = MergedLayoutStyle
    row = disclaimer_row_after(last_content_row)
    end = get_column_letter(last_col)
    ws.merge_cells(f"A{row}:{end}{row}")
    cell = ws.cell(row=row, column=1, value=PRINT_DISCLAIMER_FULL)
    cell.font = style.font(size=style.size_disclaimer, color=style.disclaimer_text)
    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ws.row_dimensions[row].height = style.disclaimer_row_height
    return row


def _write_report_banner_openpyxl(
    ws: Worksheet,
    as_of: date,
    *,
    last_col: int = PRINT_LAST_COL,
) -> int:
    """頁首：公司名 / Logo 列 / 主標題 / 英文副標 / 日期 / 空白分隔列。Logo 圖片稍後置中插入。"""
    style = MergedLayoutStyle
    end = get_column_letter(last_col)
    center = style.center()
    accent = style.fill(style.accent)
    date_fill = style.fill(style.date_bg)

    ws.merge_cells(f"A1:{end}1")
    ws["A1"] = f"{COMPANY_ZH}  /  {COMPANY_EN}"
    ws["A1"].font = style.font(size=style.size_company, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A2:{end}2")
    ws["A2"].font = style.font(size=style.size_body)
    ws.row_dimensions[2].height = style.row_logo_height

    ws.merge_cells(f"A3:{end}3")
    ws["A3"] = REPORT_TITLE_ZH
    ws["A3"].font = style.font(size=style.size_title, bold=True, color=style.text_cream)
    ws["A3"].alignment = center
    _print_fill_range(ws, 3, last_col, accent)
    ws.row_dimensions[3].height = style.row_title_height

    ws.merge_cells(f"A4:{end}4")
    ws["A4"] = REPORT_TITLE_EN
    ws["A4"].font = style.font(size=style.size_subtitle, color=style.text_cream)
    ws["A4"].alignment = center
    _print_fill_range(ws, 4, last_col, accent)
    ws.row_dimensions[4].height = style.row_subtitle_height

    ws.merge_cells(f"A5:{end}5")
    ws["A5"] = f"報價日期 / As of {as_of.strftime('%Y-%m-%d')}"
    ws["A5"].font = style.font(size=style.size_date, bold=True, color=style.date_text)
    ws["A5"].alignment = center
    _print_fill_range(ws, 5, last_col, date_fill)
    ws.row_dimensions[5].height = style.row_date_height

    ws.row_dimensions[6].height = style.row_spacer_height
    ws["A6"].font = style.font(size=style.size_body)
    return 7


def _section_heading_openpyxl(
    ws: Worksheet, row: int, zh: str, en: str | None = None, *, last_col: int = PRINT_LAST_COL
) -> int:
    style = MergedLayoutStyle
    end = get_column_letter(last_col)
    accent = style.fill(style.accent)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=f"{zh} / {en}" if en else zh)
    cell.font = style.font(size=style.size_section, bold=True, color=style.text_cream)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = style.row_section_height
    _print_fill_range(ws, row, last_col, accent)
    ws[f"A{row}"].fill = accent
    ws[f"{end}{row}"].fill = accent
    return row + 1


def _chart_image_anchors(start_row: int) -> list[str]:
    anchors: list[str] = []
    for pair in range(3):
        row = start_row + pair * CHART_ROW_STRIDE
        anchors.extend((f"A{row}", f"J{row}"))
    return anchors


def _place_chart_images(
    ws: Worksheet,
    images: list[Path],
    config: AppConfig,
    *,
    start_row: int,
) -> int:
    """獨立「遠期走勢圖」sheet：2 欄（A / J）× 3 列。"""
    anchors = _chart_image_anchors(start_row)
    for img_path, anchor in zip(images, anchors, strict=True):
        _add_anchored_image(
            ws,
            img_path,
            anchor=anchor,
            width=config.chart.image_width,
            height=config.chart.image_height,
        )
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 18)
    ws.column_dimensions["J"].width = max(ws.column_dimensions["J"].width or 0, 18)
    return start_row + 3 * CHART_ROW_STRIDE


def _place_print_chart_images(ws: Worksheet, images: list[Path], *, start_row: int) -> int:
    """合併版面走勢圖：2 欄（A / D）× 3 列，定案尺寸約 4吋 × 2.6吋。"""
    style = MergedLayoutStyle
    width = style.chart_width_px()
    height = style.chart_height_px()
    for img_path, anchor in zip(images, style.chart_anchors(start_row), strict=True):
        _add_anchored_image(ws, img_path, anchor=anchor, width=width, height=height)
    return start_row + 3 * style.chart_row_stride


def _print_bbg_row_kind(values: tuple[tuple[Any, ...], ...], r_idx: int) -> str:
    if r_idx == 0:
        return "code"
    if not values:
        return "data"
    codes = {str(v).strip().upper() for v in values[0][2:8] if v is not None and str(v).strip()}
    if codes & set(PRINT_BBG_METAL_CODES) and r_idx == 1:
        return "sub"
    return "data"


def _style_print_bbg_cell(cell: Any, *, kind: str, excel_col: int) -> None:
    style = MergedLayoutStyle
    if kind == "code" and excel_col >= 3:
        cell.fill = style.fill(style.accent)
        cell.font = style.font(size=style.size_code_header, bold=True, color=style.text_white)
        cell.alignment = style.center()
        cell.border = style.thin_border()
        return
    if kind == "code":
        cell.font = style.font(size=style.size_body)
        return
    if kind == "sub":
        cell.font = style.font(size=style.size_body)
        cell.border = style.bottom_border()
        cell.alignment = style.center()
        return
    cell.font = style.font(size=style.size_body)
    cell.border = style.thin_border()
    cell.alignment = style.center()


def _write_print_bbg_openpyxl(
    ws: Worksheet,
    values: tuple[tuple[Any, ...], ...],
    formats: tuple[tuple[str, ...], ...],
    *,
    start_row: int,
) -> int:
    """合併版面 BBG 區塊：沿用既有資料列，套用定案表頭/子表頭/數據列樣式。"""
    last_row = start_row - 1
    for r_idx, row in enumerate(values):
        fmt_row = formats[r_idx] if r_idx < len(formats) else ()
        excel_row = start_row + r_idx
        last_row = excel_row
        kind = _print_bbg_row_kind(values, r_idx)
        for c_idx, raw in enumerate(row):
            fmt = fmt_row[c_idx] if c_idx < len(fmt_row) else None
            value, num_fmt = _prepare_written_value(raw, fmt, is_header=(r_idx == 0))
            cell = ws.cell(row=excel_row, column=c_idx + 1, value=value)
            if num_fmt:
                cell.number_format = num_fmt
            elif isinstance(value, datetime):
                cell.number_format = "YYYY-MM-DD"
            elif isinstance(value, date):
                cell.number_format = "YYYY-MM-DD"
            _style_print_bbg_cell(cell, kind=kind, excel_col=c_idx + 1)
    return last_row


def _used_last_col(ws: Worksheet, start_row: int, last_row: int) -> int:
    last_col = 1
    for row in ws.iter_rows(min_row=start_row, max_row=max(last_row, start_row)):
        for cell in row:
            if cell.value is not None:
                last_col = max(last_col, cell.column)
    return last_col


def _style_print_raw_block_openpyxl(ws: Worksheet, start_row: int, last_row: int) -> None:
    """原始數據區塊：子表頭 + 數據列（置中、細框線、宋体 11pt）。"""
    if last_row < start_row:
        return
    style = MergedLayoutStyle
    last_col = _used_last_col(ws, start_row, last_row)
    for r in range(start_row, last_row + 1):
        is_header = r == start_row
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if is_header:
                cell.font = style.font(size=style.size_body)
                cell.border = style.bottom_border()
            else:
                cell.font = style.font(size=style.size_body)
                cell.border = style.thin_border()
            cell.alignment = style.center()


def _write_print_sheet_openpyxl(
    wb: Workbook,
    *,
    as_of: date,
    step2_path: Path,
    bbg_values: tuple[tuple[Any, ...], ...],
    bbg_formats: tuple[tuple[str, ...], ...],
    config: AppConfig,
    chart_images: list[Path],
) -> None:
    style = MergedLayoutStyle
    ws = wb.create_sheet(SHEET_PRINT)
    next_row = _write_report_banner_openpyxl(ws, as_of)
    header = f"{REPORT_TITLE_ZH} / {REPORT_TITLE_EN}  {as_of.strftime('%Y-%m-%d')}"

    _section_heading_openpyxl(ws, next_row, SHEET_BBG, "Bloomberg Snapshot")
    last = _write_print_bbg_openpyxl(
        ws, bbg_values, bbg_formats, start_row=next_row + 1
    )
    if last < next_row + 1:
        last = next_row
    last_content_row = last

    chart_heading = last + 2
    _section_heading_openpyxl(ws, chart_heading, SHEET_CHART, "Forward Curve")
    last_content_row = max(last_content_row, chart_heading)
    after_images = _place_print_chart_images(ws, chart_images, start_row=chart_heading + 2)
    last_content_row = max(last_content_row, after_images - 1)

    raw_heading = after_images + 1
    _section_heading_openpyxl(ws, raw_heading, PRINT_RAW_SECTION_TITLE)
    last_content_row = max(last_content_row, raw_heading)
    last_raw = _copy_raw_sheet(step2_path, ws, start_row=raw_heading + 1)
    _style_print_raw_block_openpyxl(ws, raw_heading + 1, last_raw)
    last_content_row = max(last_content_row, last_raw)

    disc_row = _write_print_disclaimer_openpyxl(ws, last_content_row=last_content_row)
    _apply_print_layout_openpyxl(
        ws,
        header=header,
        tab_color=style.accent,
        paper_size=PAPERSIZE_A3,
        header_color=style.date_text,
    )
    _set_print_area(ws, disc_row, last_col=style.last_col)
    ws.oddHeader.left.text = header_footer_run(style.date_text, COMPANY_ZH)
    ws.oddHeader.right.text = header_footer_run(style.date_text, as_of.strftime("%Y-%m-%d"))
    apply_print_data_styles(ws)
    _place_print_logo(ws, config.branding.logo_path, merge_range=PRINT_LOGO_MERGE_RANGE)


def _build_with_openpyxl(
    dest: Path,
    step2_path: Path,
    plot_df: pd.DataFrame,
    bbg_values: tuple[tuple[Any, ...], ...],
    bbg_formats: tuple[tuple[str, ...], ...],
    config: AppConfig,
    as_of: date,
) -> None:
    wb = Workbook()
    header = f"{REPORT_TITLE_ZH} {as_of.strftime('%Y-%m-%d')}"
    ws_bbg = wb.active
    ws_bbg.title = SHEET_BBG
    _write_bbg_sheet_openpyxl(ws_bbg, bbg_values, bbg_formats)
    apply_autofit_columns(ws_bbg)
    apply_center_alignment(ws_bbg)
    _apply_print_layout_openpyxl(ws_bbg, header=f"{header} · {SHEET_BBG}", tab_color=COLOR_NAVY)

    tmp_dir, images = _render_matplotlib_pngs(plot_df, config)
    try:
        ws_chart = wb.create_sheet(SHEET_CHART)
        ws_chart["A1"] = "LME 遠期曲線（cash date → +{} 個月）".format(config.chart.forward_months)
        ws_chart["A1"].font = Font(bold=True, size=14, color=COLOR_NAVY)
        logger.debug("%s：%s", SOURCE_LABEL, _win_path(step2_path))
        _place_chart_images(ws_chart, images, config, start_row=4)
        apply_autofit_columns(ws_chart)
        _apply_print_layout_openpyxl(
            ws_chart, header=f"{header} · {SHEET_CHART}", tab_color=COLOR_GOLD
        )

        ws_raw = wb.create_sheet(SHEET_RAW)
        _copy_raw_sheet(step2_path, ws_raw)
        apply_autofit_columns(ws_raw)
        _apply_print_layout_openpyxl(
            ws_raw, header=f"{header} · {SHEET_RAW}", tab_color=COLOR_MUTED
        )

        _write_print_sheet_openpyxl(
            wb,
            as_of=as_of,
            step2_path=step2_path,
            bbg_values=bbg_values,
            bbg_formats=bbg_formats,
            config=config,
            chart_images=images,
        )

        try:
            wb.save(dest)
        except Exception as exc:
            raise ReportBuildError(f"儲存報告失敗（{dest}）：{exc}") from exc
    finally:
        wb.close()
        shutil.rmtree(tmp_dir, ignore_errors=True)


def _configure_cjk_matplotlib() -> None:
    """讓圖表標題的中文在 Windows（微軟正黑體 / 雅黑）能正常顯示。"""
    import matplotlib
    from matplotlib import font_manager

    matplotlib.rcParams["axes.unicode_minus"] = False
    candidates = [
        Path(r"C:\Windows\Fonts\msjh.ttc"),
        Path(r"C:\Windows\Fonts\msjhbd.ttc"),
        Path(r"C:\Windows\Fonts\msyh.ttc"),
        Path(r"C:\Windows\Fonts\simhei.ttf"),
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc"),
        Path("/usr/share/fonts/opentype/noto/NotoSansCJKtc-Regular.otf"),
    ]
    for font_path in candidates:
        if font_path.is_file():
            try:
                font_manager.fontManager.addfont(str(font_path))
                name = font_manager.FontProperties(fname=str(font_path)).get_name()
                matplotlib.rcParams["font.sans-serif"] = [name, "DejaVu Sans"]
                matplotlib.rcParams["font.family"] = "sans-serif"
                logger.info("matplotlib 使用中文字型：%s", font_path)
                return
            except Exception as exc:
                logger.debug("載入字型失敗 %s：%s", font_path, exc)
    logger.warning(
        "找不到中文字型，遠期曲線標題可能出現方塊。Windows 請確認已安裝微軟正黑體。"
    )


def _render_matplotlib_pngs(plot_df: pd.DataFrame, config: AppConfig) -> tuple[Path, list[Path]]:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    _configure_cjk_matplotlib()
    tmp_dir = Path(tempfile.mkdtemp(prefix="lme_charts_"))
    images: list[Path] = []
    for code, title in SIX_SERIES.items():
        series_df = plot_df.dropna(subset=[code])[[DATE_COL, code]]
        if series_df.empty:
            logger.warning("品種 %s 在圖表視窗內沒有非空資料，仍會放空圖說明", code)
        fig, ax = plt.subplots(figsize=(6.4, 3.6), dpi=120)
        fig.patch.set_facecolor("white")
        ax.set_facecolor("white")
        if not series_df.empty:
            y = series_df[code].map(lambda v: value_from_stored_number(v) if pd.notna(v) else v)
            ax.plot(
                series_df[DATE_COL],
                y,
                color=_css(COLOR_NAVY),
                linewidth=1.7,
            )
        ax.set_title(f"{title} 遠期曲線", color=_css(COLOR_NAVY), fontweight="bold")
        ax.set_xlabel("到期日 Prompt", color=_css(COLOR_MUTED))
        ax.set_ylabel("價格", color=_css(COLOR_MUTED))
        ax.tick_params(colors=_css(COLOR_INK))
        ax.grid(True, color="#E5E7EB", alpha=0.95)
        for spine in ax.spines.values():
            spine.set_color(_css(COLOR_GOLD))
            spine.set_linewidth(0.8)
        fig.autofmt_xdate()
        fig.tight_layout()
        img_path = tmp_dir / f"{code}.png"
        fig.savefig(img_path, dpi=120, facecolor=fig.get_facecolor())
        plt.close(fig)
        images.append(img_path)
    return tmp_dir, images


def _build_with_xlsxwriter(
    dest: Path,
    step2_path: Path,
    curve_df: pd.DataFrame,
    plot_df: pd.DataFrame,
    bbg_values: tuple[tuple[Any, ...], ...],
    bbg_formats: tuple[tuple[str, ...], ...],
    config: AppConfig,
    as_of: date,
) -> None:
    import xlsxwriter

    tmp_dir: Path | None = None
    try:
        workbook = xlsxwriter.Workbook(str(dest))
    except Exception as exc:
        raise ReportBuildError(f"無法建立 xlsxwriter 工作簿（{dest}）：{exc}") from exc

    try:
        header = f"{REPORT_TITLE_ZH} {as_of.strftime('%Y-%m-%d')}"
        _write_bbg_sheet_xlsxwriter(workbook, bbg_values, bbg_formats, header=header)
        _write_xlsxwriter_charts(workbook, plot_df, config, header=header, step2_path=step2_path)
        _write_raw_sheet_xlsxwriter(workbook, step2_path, curve_df, header=header)
        tmp_dir, images = _render_matplotlib_pngs(plot_df, config)
        _write_print_sheet_xlsxwriter(
            workbook,
            as_of=as_of,
            step2_path=step2_path,
            curve_df=curve_df,
            bbg_values=bbg_values,
            bbg_formats=bbg_formats,
            config=config,
            chart_images=images,
            header=header,
        )
        workbook.close()
    except ReportBuildError:
        try:
            workbook.close()
        except Exception:
            pass
        raise
    except Exception as exc:
        try:
            workbook.close()
        except Exception:
            pass
        raise ReportBuildError(f"xlsxwriter 產生報告失敗：{exc}") from exc
    finally:
        if tmp_dir:
            shutil.rmtree(tmp_dir, ignore_errors=True)


def _apply_print_layout_xlsxwriter(
    ws: Any, *, header: str, paper: int = 9, header_color: str = COLOR_NAVY
) -> None:
    ws.set_landscape()
    ws.set_paper(paper)
    ws.fit_to_pages(1, 0)
    ws.set_header(f"&C{header_footer_run(header_color, header)}")
    ws.set_footer(f"&C{PRINT_FOOTER}")
    ws.center_horizontally()
    ws.set_margins(left=0.4, right=0.4, top=0.75, bottom=0.7)


def _write_bbg_rows_xlsxwriter(
    ws: Any,
    workbook: Any,
    values: tuple[tuple[Any, ...], ...],
    formats: tuple[tuple[str, ...], ...],
    *,
    start_row: int = 0,
    header_fill_color: str = COLOR_NAVY,
    center_align: bool = False,
    widths: _ColWidthAcc | None = None,
    apply_widths: bool = True,
) -> int:
    align = {"align": "center", "valign": "vcenter"} if center_align else {}
    header_fmt = workbook.add_format(
        {
            "bold": True,
            "bg_color": _css(header_fill_color),
            "font_color": _css(COLOR_WHITE),
            **align,
        }
    )
    date_fmt = workbook.add_format({"num_format": "yyyy-mm-dd", **align})
    price_fmt = workbook.add_format({"num_format": NUMBER_FORMAT_2DP, **align})
    default_fmt = workbook.add_format(align) if center_align else None
    acc = widths if widths is not None else _ColWidthAcc()
    last = start_row - 1
    for r_idx, row in enumerate(values):
        fmt_row = formats[r_idx] if r_idx < len(formats) else ()
        excel_row = start_row + r_idx
        last = excel_row
        for c_idx, raw in enumerate(row):
            fmt = fmt_row[c_idx] if c_idx < len(fmt_row) else None
            value, num_fmt = _prepare_written_value(raw, fmt, is_header=(r_idx == 0))
            cell_fmt = header_fmt if r_idx == 0 else None
            if num_fmt == "YYYY-MM-DD" and r_idx != 0:
                if isinstance(value, date) and not isinstance(value, datetime):
                    value = datetime(value.year, value.month, value.day)
                cell_fmt = date_fmt
            elif num_fmt == NUMBER_FORMAT_2DP:
                cell_fmt = price_fmt
            if isinstance(value, date) and not isinstance(value, datetime) and r_idx != 0:
                value = datetime(value.year, value.month, value.day)
            if cell_fmt is None:
                cell_fmt = default_fmt
            ws.write(excel_row, c_idx, value, cell_fmt)
            acc.add(c_idx, value, number_format=num_fmt)
    if apply_widths:
        acc.apply_xlsxwriter(ws)
    return last


def _write_bbg_sheet_xlsxwriter(
    workbook: Any,
    values: tuple[tuple[Any, ...], ...],
    formats: tuple[tuple[str, ...], ...],
    *,
    header: str,
) -> None:
    ws = workbook.add_worksheet(SHEET_BBG)
    _apply_print_layout_xlsxwriter(ws, header=f"{header} · {SHEET_BBG}")
    ws.set_tab_color(_css(COLOR_NAVY))
    _write_bbg_rows_xlsxwriter(
        ws, workbook, values, formats, start_row=0, center_align=True
    )


def _write_xlsxwriter_charts(
    workbook: Any,
    plot_df: pd.DataFrame,
    config: AppConfig,
    *,
    header: str,
    step2_path: Path,
) -> None:
    data_ws = workbook.add_worksheet(SHEET_CHART_DATA)
    chart_ws = workbook.add_worksheet(SHEET_CHART)
    data_ws.hide()
    title_fmt = workbook.add_format(
        {"bold": True, "font_size": 14, "font_color": _css(COLOR_NAVY)}
    )
    chart_ws.write(0, 0, f"LME 遠期曲線（cash date → +{config.chart.forward_months} 個月）", title_fmt)
    logger.debug("%s：%s", SOURCE_LABEL, _win_path(step2_path))
    _apply_print_layout_xlsxwriter(chart_ws, header=f"{header} · {SHEET_CHART}")
    chart_ws.set_tab_color(_css(COLOR_GOLD))

    date_fmt = workbook.add_format({"num_format": "yyyy-mm-dd"})
    price_fmt = workbook.add_format({"num_format": NUMBER_FORMAT_2DP})
    positions = [
        ("A4", "J4"),
        ("A23", "J23"),
        ("A42", "J42"),
    ]
    data_widths = _ColWidthAcc()
    col = 0
    for idx, (code, title) in enumerate(SIX_SERIES.items()):
        series_df = plot_df.dropna(subset=[code])[[DATE_COL, code]].reset_index(drop=True)
        data_ws.write(0, col, f"{code}_date")
        data_ws.write(0, col + 1, f"{code}_px")
        data_widths.add(col, f"{code}_date")
        data_widths.add(col + 1, f"{code}_px")
        for r, rec in series_df.iterrows():
            ts = rec[DATE_COL]
            if isinstance(ts, pd.Timestamp):
                ts = ts.to_pydatetime()
            px = value_from_stored_number(rec[code])
            data_ws.write_datetime(r + 1, col, ts, date_fmt)
            data_ws.write_number(r + 1, col + 1, float(px), price_fmt)
            data_widths.add(col, ts, number_format="yyyy-mm-dd")
            data_widths.add(col + 1, px, number_format=NUMBER_FORMAT_2DP)

        chart = workbook.add_chart({"type": "line"})
        n = len(series_df)
        if n:
            chart.add_series(
                {
                    "name": title,
                    "categories": [SHEET_CHART_DATA, 1, col, n, col],
                    "values": [SHEET_CHART_DATA, 1, col + 1, n, col + 1],
                    "line": {"color": _css(COLOR_NAVY), "width": 1.5},
                }
            )
        chart.set_title({"name": f"{title} 遠期曲線"})
        chart.set_x_axis({"name": "到期日 Prompt", "num_format": "yyyy-mm-dd"})
        chart.set_y_axis({"name": "價格"})
        chart.set_size({"width": config.chart.image_width, "height": config.chart.image_height})
        chart.set_legend({"none": True})
        row_pair, col_slot = divmod(idx, 2)
        anchor = positions[row_pair][col_slot]
        chart_ws.insert_chart(anchor, chart)
        col += 2
    data_widths.apply_xlsxwriter(data_ws)
    chart_widths = _ColWidthAcc()
    chart_widths.add(0, f"LME 遠期曲線（cash date → +{config.chart.forward_months} 個月）")
    chart_widths.apply_xlsxwriter(chart_ws)


def _write_raw_sheet_xlsxwriter(
    workbook: Any,
    step2_path: Path,
    curve_df: pd.DataFrame,
    *,
    header: str,
    worksheet: Any | None = None,
    start_row: int = 0,
    center_align: bool = False,
    widths: _ColWidthAcc | None = None,
    apply_widths: bool | None = None,
    print_style: bool = False,
) -> int:
    ws = worksheet or workbook.add_worksheet(SHEET_RAW)
    if worksheet is None:
        _apply_print_layout_xlsxwriter(ws, header=f"{header} · {SHEET_RAW}")
        ws.set_tab_color(_css(COLOR_MUTED))
    style = MergedLayoutStyle
    if print_style:
        base = {
            "font_name": style.font_name,
            "font_size": style.size_body,
            "align": "center",
            "valign": "vcenter",
        }
        header_fmt = workbook.add_format({**base, "bottom": 1})
        date_fmt = workbook.add_format({**base, "num_format": "yyyy-mm-dd", "border": 1})
        price_fmt = workbook.add_format({**base, "num_format": NUMBER_FORMAT_2DP, "border": 1})
        default_fmt = workbook.add_format({**base, "border": 1})
    else:
        align = {"align": "center", "valign": "vcenter"} if center_align else {}
        header_fmt = workbook.add_format({"bold": True, **align})
        date_fmt = workbook.add_format({"num_format": "yyyy-mm-dd", **align})
        price_fmt = workbook.add_format({"num_format": NUMBER_FORMAT_2DP, **align})
        default_fmt = workbook.add_format(align) if center_align else None
    acc = widths if widths is not None else _ColWidthAcc()
    own_sheet = worksheet is None
    if apply_widths is None:
        apply_widths = own_sheet
    last = start_row - 1
    try:
        src_wb = load_workbook(step2_path, data_only=True)
        src_ws = src_wb.active
        for row in src_ws.iter_rows():
            for cell in row:
                value = cell.value
                if _is_excel_formula(value):
                    logger.debug("原始數據略過公式儲存格 %s", cell.coordinate)
                    continue
                if value is None:
                    continue
                r, c = cell.row - 1 + start_row, cell.column - 1
                last = max(last, r)
                is_header = cell.row == 1
                if not is_header:
                    value = value_from_stored_number(value)
                fmt = header_fmt if is_header else default_fmt
                num_fmt: str | None = None
                if isinstance(value, datetime):
                    ws.write_datetime(r, c, value, date_fmt)
                    num_fmt = "yyyy-mm-dd"
                elif isinstance(value, date):
                    ws.write_datetime(r, c, datetime(value.year, value.month, value.day), date_fmt)
                    num_fmt = "yyyy-mm-dd"
                elif _is_number_like(value):
                    ws.write_number(r, c, float(value), price_fmt)
                    num_fmt = NUMBER_FORMAT_2DP
                else:
                    ws.write(r, c, value, fmt)
                acc.add(c, value, number_format=num_fmt)
        src_wb.close()
        if apply_widths:
            acc.apply_xlsxwriter(ws)
        return last
    except Exception as exc:
        logger.warning("xlsxwriter 逐格複製原始數據失敗，改用 pandas：%s", exc)

    for c_idx, col_name in enumerate(curve_df.columns):
        ws.write(start_row, c_idx, col_name, header_fmt)
        acc.add(c_idx, col_name)
        last = start_row
    for r_idx, rec in curve_df.iterrows():
        excel_row = int(r_idx) + 1 + start_row
        last = max(last, excel_row)
        for c_idx, col_name in enumerate(curve_df.columns):
            value = rec[col_name]
            if pd.isna(value):
                continue
            if col_name == DATE_COL and isinstance(value, pd.Timestamp):
                ws.write_datetime(excel_row, c_idx, value.to_pydatetime(), date_fmt)
                acc.add(c_idx, value.to_pydatetime(), number_format="yyyy-mm-dd")
            else:
                value = value_from_stored_number(value)
                if _is_number_like(value):
                    ws.write_number(excel_row, c_idx, float(value), price_fmt)
                    acc.add(c_idx, value, number_format=NUMBER_FORMAT_2DP)
                else:
                    ws.write(excel_row, c_idx, value, default_fmt)
                    acc.add(c_idx, value)
    if apply_widths:
        acc.apply_xlsxwriter(ws)
    return last


def _write_print_bbg_xlsxwriter(
    ws: Any,
    workbook: Any,
    values: tuple[tuple[Any, ...], ...],
    formats: tuple[tuple[str, ...], ...],
    *,
    start_row: int,
    widths: _ColWidthAcc | None = None,
) -> int:
    style = MergedLayoutStyle
    body = {
        "font_name": style.font_name,
        "font_size": style.size_body,
        "align": "center",
        "valign": "vcenter",
    }
    code_fmt = style.xw_format(
        workbook,
        bold=True,
        font_size=style.size_code_header,
        font_color=_css(style.text_white),
        bg_color=_css(style.accent),
        align="center",
        valign="vcenter",
        border=1,
    )
    code_plain = style.xw_format(workbook, font_size=style.size_body)
    sub_fmt = style.xw_format(workbook, **body, bottom=1)
    data_fmt = style.xw_format(workbook, **body, border=1)
    date_fmt = style.xw_format(workbook, **body, border=1, num_format="yyyy-mm-dd")
    price_fmt = style.xw_format(workbook, **body, border=1, num_format=NUMBER_FORMAT_2DP)
    acc = widths if widths is not None else _ColWidthAcc()
    last = start_row - 1
    for r_idx, row in enumerate(values):
        fmt_row = formats[r_idx] if r_idx < len(formats) else ()
        excel_row = start_row + r_idx
        last = excel_row
        kind = _print_bbg_row_kind(values, r_idx)
        for c_idx, raw in enumerate(row):
            fmt = fmt_row[c_idx] if c_idx < len(fmt_row) else None
            value, num_fmt = _prepare_written_value(raw, fmt, is_header=(r_idx == 0))
            if kind == "code":
                cell_fmt = code_fmt if c_idx + 1 >= 3 else code_plain
            elif kind == "sub":
                cell_fmt = sub_fmt
            elif num_fmt == "YYYY-MM-DD":
                cell_fmt = date_fmt
                if isinstance(value, date) and not isinstance(value, datetime):
                    value = datetime(value.year, value.month, value.day)
            elif num_fmt == NUMBER_FORMAT_2DP:
                cell_fmt = price_fmt
            else:
                cell_fmt = data_fmt
            if isinstance(value, date) and not isinstance(value, datetime) and kind == "data":
                value = datetime(value.year, value.month, value.day)
            ws.write(excel_row, c_idx, value, cell_fmt)
            acc.add(c_idx, value, number_format=num_fmt)
    return last


def _apply_merged_layout_autofit_xlsxwriter(ws: Any, acc: _ColWidthAcc) -> list[float]:
    """A:H 一律自動欄寬，回傳各欄字元寬（供 Logo 置中）。"""
    widths: list[float] = []
    for col in range(PRINT_LAST_COL):
        fitted = acc.fitted_width(col)
        if fitted is None:
            widths.append(EXCEL_DEFAULT_COLUMN_WIDTH)
            continue
        ws.set_column(col, col, fitted)
        widths.append(fitted)
    return widths


def _write_print_sheet_xlsxwriter(
    workbook: Any,
    *,
    as_of: date,
    step2_path: Path,
    curve_df: pd.DataFrame,
    bbg_values: tuple[tuple[Any, ...], ...],
    bbg_formats: tuple[tuple[str, ...], ...],
    config: AppConfig,
    chart_images: list[Path],
    header: str,
) -> None:
    style = MergedLayoutStyle
    ws = workbook.add_worksheet(SHEET_PRINT)
    last_col = style.last_col - 1
    company = style.xw_format(
        workbook, bold=True, font_size=style.size_company, align="left", valign="vcenter"
    )
    title = style.xw_format(
        workbook,
        bold=True,
        font_size=style.size_title,
        font_color=_css(style.text_cream),
        bg_color=_css(style.accent),
        align="center",
        valign="vcenter",
    )
    subtitle = style.xw_format(
        workbook,
        font_size=style.size_subtitle,
        font_color=_css(style.text_cream),
        bg_color=_css(style.accent),
        align="center",
        valign="vcenter",
    )
    date_fmt = style.xw_format(
        workbook,
        bold=True,
        font_size=style.size_date,
        font_color=_css(style.date_text),
        bg_color=_css(style.date_bg),
        align="center",
        valign="vcenter",
    )
    blank = style.xw_format(workbook)
    section = style.xw_format(
        workbook,
        bold=True,
        font_size=style.size_section,
        font_color=_css(style.text_cream),
        bg_color=_css(style.accent),
        align="left",
        valign="vcenter",
    )

    ws.merge_range(0, 0, 0, last_col, f"{COMPANY_ZH}  /  {COMPANY_EN}", company)
    ws.merge_range(1, 0, 1, last_col, "", blank)
    ws.set_row(1, style.row_logo_height)
    ws.merge_range(2, 0, 2, last_col, REPORT_TITLE_ZH, title)
    ws.set_row(2, style.row_title_height)
    ws.merge_range(3, 0, 3, last_col, REPORT_TITLE_EN, subtitle)
    ws.set_row(3, style.row_subtitle_height)
    ws.merge_range(
        4, 0, 4, last_col, f"報價日期 / As of {as_of.strftime('%Y-%m-%d')}", date_fmt
    )
    ws.set_row(4, style.row_date_height)
    ws.set_row(5, style.row_spacer_height)

    print_widths = _ColWidthAcc()
    bbg_heading = 6
    ws.merge_range(bbg_heading, 0, bbg_heading, last_col, f"{SHEET_BBG} / Bloomberg Snapshot", section)
    ws.set_row(bbg_heading, style.row_section_height)
    last = _write_print_bbg_xlsxwriter(
        ws, workbook, bbg_values, bbg_formats, start_row=bbg_heading + 1, widths=print_widths
    )
    if last < bbg_heading + 1:
        last = bbg_heading
    last_content_row = last

    chart_heading = last + 2
    ws.merge_range(
        chart_heading, 0, chart_heading, last_col, f"{SHEET_CHART} / Forward Curve", section
    )
    ws.set_row(chart_heading, style.row_section_height)
    last_content_row = max(last_content_row, chart_heading)
    img_start = chart_heading + 2
    native_w = 6.4 * 120
    native_h = 3.6 * 120
    x_scale = style.chart_width_px() / native_w
    y_scale = style.chart_height_px() / native_h
    anchors = style.chart_anchors(img_start + 1)
    for img_path, anchor in zip(chart_images, anchors, strict=True):
        ws.insert_image(anchor, str(img_path), {"x_scale": x_scale, "y_scale": y_scale})
    after_images = img_start + 3 * style.chart_row_stride
    last_content_row = max(last_content_row, after_images - 1)

    raw_heading = after_images + 1
    ws.merge_range(raw_heading, 0, raw_heading, last_col, PRINT_RAW_SECTION_TITLE, section)
    ws.set_row(raw_heading, style.row_section_height)
    last_content_row = max(last_content_row, raw_heading)
    last_raw = _write_raw_sheet_xlsxwriter(
        workbook,
        step2_path,
        curve_df,
        header=header,
        worksheet=ws,
        start_row=raw_heading + 1,
        center_align=True,
        widths=print_widths,
        apply_widths=False,
        print_style=True,
    )
    last_content_row = max(last_content_row, last_raw)
    disc_row = disclaimer_row_after(last_content_row)
    disc_fmt = style.xw_format(
        workbook,
        font_size=style.size_disclaimer,
        font_color=_css(style.disclaimer_text),
        text_wrap=True,
        valign="top",
        align="left",
    )
    ws.merge_range(disc_row, 0, disc_row, last_col, PRINT_DISCLAIMER_FULL, disc_fmt)
    ws.set_row(disc_row, style.disclaimer_row_height)
    _apply_print_layout_xlsxwriter(
        ws, header=header, paper=PAPERSIZE_A3, header_color=style.date_text
    )
    ws.set_header(
        f"&L{header_footer_run(style.date_text, COMPANY_ZH)}"
        f"&C{header_footer_run(style.date_text, header)}"
        f"&R{header_footer_run(style.date_text, as_of.strftime('%Y-%m-%d'))}"
    )
    ws.set_footer(f"&C{PRINT_FOOTER}")
    ws.set_tab_color(_css(style.accent))
    ws.print_area(0, 0, disc_row, last_col)
    col_widths = _apply_merged_layout_autofit_xlsxwriter(ws, print_widths)
    _place_print_logo_xlsxwriter(ws, config.branding.logo_path, col_widths_chars=col_widths)


def merged_layout_disclaimer_row(*, bbg_rows: int, raw_rows: int) -> int:
    """合併版面免責聲明列號（1-based）。bbg_rows / raw_rows 含表頭。"""
    bbg_heading = 7
    bbg_end = bbg_heading + bbg_rows
    chart_heading = bbg_end + 2
    img_start = chart_heading + 2
    after_images = img_start + 3 * PRINT_CHART_ROW_STRIDE
    raw_heading = after_images + 1
    last_raw = raw_heading + raw_rows
    last_content_row = max(after_images - 1, raw_heading, last_raw)
    return disclaimer_row_after(last_content_row)


def dataframe_preview_rows(df: pd.DataFrame, limit: int = 5) -> list[list[Any]]:
    """測試輔助：把 DataFrame 前幾列轉成純 Python。"""
    rows = []
    for row in dataframe_to_rows(df.head(limit), index=False, header=True):
        rows.append(list(row))
    return rows
