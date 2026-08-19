"""用 Bloomberg Desktop API（blpapi）取快照，不經 Excel 外掛。

仍需要本機 Bloomberg Terminal **已登入且未鎖定**。程式不能自動解鎖
（那是 Bloomberg 安全機制，也違反使用條款）。

若 Terminal 已鎖：請在 Bloomberg 視窗手動解鎖後再跑。
無人值守請改走 Bloomberg Data License / B-PIPE（需另外簽約）。
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from lme_daily.bbg_fetch import RangeFormats, RangeValues, parse_bdp_formula
from lme_daily.config import AppConfig
from lme_daily.exceptions import ExcelComError

logger = logging.getLogger(__name__)


def fetch_bloomberg_snapshot_blpapi(config: AppConfig) -> tuple[RangeValues, RangeFormats]:
    try:
        import blpapi  # type: ignore
    except ImportError as exc:
        raise ExcelComError(
            "bloomberg.source=blpapi 需要套件 blpapi。請執行：pip install blpapi\n"
            "Desktop API 仍須 Terminal 已登入未鎖定；無法自動解鎖。"
        ) from exc

    formula_grid = _read_formula_grid(
        config.paths.bbg_workbook,
        config.bloomberg.bbg_sheet_name,
        config.bloomberg.copy_range,
    )
    pairs = _collect_bdp_pairs(formula_grid, config)
    if not pairs:
        raise ExcelComError(
            "blpapi 模式找不到 BDP 公式，也沒有 bloomberg.securities。"
            "請在 LME BBG WORKBOOK 的 copy_range 使用 =BDP(\"ticker\",\"FIELD\")，"
            "或在 config.yaml 列出 securities / fields。"
        )

    logger.info("blpapi 連線 %s:%s，請求 %d 檔", config.bloomberg.host, config.bloomberg.port, len(pairs))
    values_map = _refdata_request(blpapi, config, pairs)
    grid = _overlay_values(formula_grid, values_map, config)
    formats = tuple(tuple("General" for _ in row) for row in grid)
    logger.info("blpapi 已填入 %d 列", len(grid))
    return grid, formats


def _read_formula_grid(path: Path, sheet_name: str, copy_range: str) -> list[list[Any]]:
    if not path.is_file():
        raise ExcelComError(f"BBG 工作簿不存在：{path}")
    wb = load_workbook(path, data_only=False, read_only=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise ExcelComError(
                f"找不到工作表 {sheet_name!r}。實際：{wb.sheetnames}"
            )
        ws = wb[sheet_name]
        cells = ws[copy_range]
        rows: list[list[Any]] = []
        if not isinstance(cells, tuple):
            cells = ((cells,),)
        for row in cells:
            if not isinstance(row, tuple):
                row = (row,)
            rows.append([cell.value for cell in row])
        return rows
    finally:
        wb.close()


def _collect_bdp_pairs(formula_grid: list[list[Any]], config: AppConfig) -> list[tuple[str, str]]:
    found: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for row in formula_grid:
        for cell in row:
            parsed = parse_bdp_formula(cell)
            if parsed and parsed not in seen:
                seen.add(parsed)
                found.append(parsed)
    if found:
        return found
    securities = list(config.bloomberg.securities)
    fields = list(config.bloomberg.fields) or ["PX_LAST"]
    for sec in securities:
        for field in fields:
            pair = (sec, field)
            if pair not in seen:
                seen.add(pair)
                found.append(pair)
    return found


def _overlay_values(
    formula_grid: list[list[Any]],
    values_map: dict[tuple[str, str], Any],
    config: AppConfig,
) -> RangeValues:
    has_bdp = any(parse_bdp_formula(c) for row in formula_grid for c in row)
    if has_bdp:
        out_rows = []
        for row in formula_grid:
            new_row = []
            for cell in row:
                parsed = parse_bdp_formula(cell)
                if parsed:
                    new_row.append(values_map.get(parsed))
                else:
                    new_row.append(cell)
            out_rows.append(tuple(new_row))
        return tuple(out_rows)

    # 無公式時：第一列欄位名，其後每列一個 security
    fields = list(config.bloomberg.fields) or ["PX_LAST"]
    securities = list(config.bloomberg.securities)
    header = tuple(["Security", *fields])
    data_rows = []
    for sec in securities:
        data_rows.append(tuple([sec, *[values_map.get((sec, f)) for f in fields]]))
    return (header, *data_rows)


def _refdata_request(blpapi: Any, config: AppConfig, pairs: list[tuple[str, str]]) -> dict[tuple[str, str], Any]:
    options = blpapi.SessionOptions()
    options.setServerHost(config.bloomberg.host)
    options.setServerPort(config.bloomberg.port)
    session = blpapi.Session(options)
    if not session.start():
        raise ExcelComError(
            "blpapi 無法連上 Desktop API。請確認 Bloomberg Terminal 已開啟、已登入、未鎖定。"
            "程式不能自動解鎖 Terminal。"
        )
    try:
        if not session.openService("//blp/refdata"):
            raise ExcelComError("無法開啟 //blp/refdata。Terminal 可能已鎖定或授權不足。")
        service = session.getService("//blp/refdata")
        request = service.createRequest("ReferenceDataRequest")
        securities = sorted({sec for sec, _ in pairs})
        fields = sorted({field for _, field in pairs})
        for sec in securities:
            request.getElement("securities").appendValue(sec)
        for field in fields:
            request.getElement("fields").appendValue(field)
        session.sendRequest(request)

        out: dict[tuple[str, str], Any] = {}
        while True:
            event = session.nextEvent(5000)
            for msg in event:
                if msg.hasElement("responseError"):
                    raise ExcelComError(f"blpapi responseError：{msg}")
                if not msg.hasElement("securityData"):
                    continue
                sec_data = msg.getElement("securityData")
                for i in range(sec_data.numValues()):
                    item = sec_data.getValueAsElement(i)
                    sec = item.getElementAsString("security")
                    if item.hasElement("securityError"):
                        logger.warning("securityError %s：%s", sec, item.getElement("securityError"))
                        continue
                    field_data = item.getElement("fieldData")
                    for field in fields:
                        if field_data.hasElement(field):
                            out[(sec, field)] = _blp_element_value(field_data.getElement(field))
            if event.eventType() == blpapi.Event.RESPONSE:
                break
        return out
    finally:
        try:
            session.stop()
        except Exception:
            pass


def _blp_element_value(element: Any) -> Any:
    try:
        return element.getValue()
    except Exception:
        try:
            return element.getValueAsString()
        except Exception:
            return None
