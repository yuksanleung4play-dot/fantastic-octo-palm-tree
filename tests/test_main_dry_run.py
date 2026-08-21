"""dry-run 串接：不算 Excel COM。"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
import yaml

from lme_daily.config import load_config
from lme_daily.main import run
from lme_daily.exceptions import LMEAutomationError
from lme_daily.report_builder import SHEET_BBG, SHEET_CHART, SHEET_PRINT, SHEET_RAW
from tests.test_report_builder import _make_step2


def test_dry_run_returns_none(tmp_path: Path):
    work = tmp_path / "work"
    work.mkdir()
    payload = {
        "paths": {
            "working_dir": str(work),
            "ref_workbook_name": "ref.xlsm",
            "bbg_workbook_name": "bbg.xlsx",
            "output_prefix": "LME每日報價",
        },
        "vba": {"macro_name": "RunDailyLME", "use_param_injection": True},
        "bloomberg": {
            "copy_range": "B3:I10",
            "bbg_sheet_name": "Promt date",
            "refresh_wait_seconds": 1,
        },
        "chart": {"forward_months": 27, "engine": "matplotlib"},
        "logging": {"level": "INFO", "file": ""},
    }
    cfg = tmp_path / "config.yaml"
    cfg.write_text(yaml.safe_dump(payload, allow_unicode=True), encoding="utf-8")
    config = load_config(cfg)
    assert run(config, as_of=date(2026, 8, 19), dry_run=True, skip_vba=False, skip_bbg=False) is None


def test_skip_vba_and_bbg_builds_report(tmp_path: Path):
    work = tmp_path / "work"
    work.mkdir()
    payload = {
        "paths": {
            "working_dir": str(work),
            "ref_workbook_name": "ref.xlsm",
            "bbg_workbook_name": "bbg.xlsx",
            "output_prefix": "LME每日報價",
        },
        "vba": {"macro_name": "RunDailyLME", "use_param_injection": True},
        "bloomberg": {
            "copy_range": "B3:I10",
            "bbg_sheet_name": "Promt date",
            "refresh_wait_seconds": 1,
        },
        "chart": {"forward_months": 27, "engine": "matplotlib"},
        "logging": {"level": "INFO", "file": ""},
    }
    cfg = tmp_path / "config.yaml"
    cfg.write_text(yaml.safe_dump(payload, allow_unicode=True), encoding="utf-8")
    _make_step2(work / "20260819.xlsx")
    config = load_config(cfg)
    dest = run(
        config,
        as_of=date(2026, 8, 19),
        dry_run=False,
        skip_vba=True,
        skip_bbg=True,
    )
    assert dest is not None
    assert dest == work / "20260819" / "LME每日報價20260819.xlsx"
    assert dest.is_file()
    assert (work / "20260819" / "20260819.xlsx").is_file()
    assert (work / "20260819" / "lme_daily.log").is_file()
    assert not (work / "20260819" / "LME每日報價20260819.pdf").exists()
    assert list(dest.parent.glob("*.pdf")) == []
    from openpyxl import load_workbook

    wb = load_workbook(dest)
    assert {SHEET_BBG, SHEET_CHART, SHEET_RAW, SHEET_PRINT} <= set(wb.sheetnames)
    wb.close()


def test_skip_vba_missing_step2_stops(tmp_path: Path):
    work = tmp_path / "work"
    work.mkdir()
    payload = {
        "paths": {
            "working_dir": str(work),
            "ref_workbook_name": "ref.xlsm",
            "bbg_workbook_name": "bbg.xlsx",
            "output_prefix": "LME每日報價",
        },
        "vba": {"macro_name": "RunDailyLME", "use_param_injection": True},
        "bloomberg": {
            "copy_range": "B3:I10",
            "bbg_sheet_name": "Promt date",
            "refresh_wait_seconds": 1,
        },
        "chart": {"forward_months": 27, "engine": "matplotlib"},
        "logging": {"level": "INFO", "file": ""},
    }
    cfg = tmp_path / "config.yaml"
    cfg.write_text(yaml.safe_dump(payload, allow_unicode=True), encoding="utf-8")
    config = load_config(cfg)
    with pytest.raises(LMEAutomationError, match="找不到 VBA 中繼檔"):
        run(
            config,
            as_of=date(2026, 8, 19),
            dry_run=False,
            skip_vba=True,
            skip_bbg=True,
        )


def test_bbg_fetch_runs_before_vba_and_injects_bbg_3m(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    work = tmp_path / "work"
    work.mkdir()
    (work / "ref.xlsm").write_bytes(b"fake")
    (work / "bbg.xlsx").write_bytes(b"fake")
    payload = {
        "paths": {
            "working_dir": str(work),
            "ref_workbook_name": "ref.xlsm",
            "bbg_workbook_name": "bbg.xlsx",
            "output_prefix": "LME每日報價",
        },
        "vba": {"macro_name": "RunDailyLME", "use_param_injection": True},
        "bloomberg": {
            "copy_range": "B3:I10",
            "bbg_sheet_name": "Promt date",
            "prompt_date_cell": "B4",
            "refresh_wait_seconds": 1,
        },
        "chart": {"forward_months": 27, "engine": "matplotlib"},
        "logging": {"level": "INFO", "file": ""},
    }
    cfg = tmp_path / "config.yaml"
    cfg.write_text(yaml.safe_dump(payload, allow_unicode=True), encoding="utf-8")
    config = load_config(cfg)

    order: list[str] = []
    captured: dict[str, str] = {}

    def fake_fetch(_cfg, *, fallback_3m=None):
        order.append("bbg")
        return ((("Metal",),), (("General",),), "20261201")

    def fake_vba(_cfg, *, as_of, prev_date, three_m_date):
        order.append("vba")
        captured["prev"] = prev_date
        captured["three_m"] = three_m_date
        dest = _cfg.step2_workbook_path(as_of)
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(b"xlsx")
        return dest

    def fake_report(*_args, **_kwargs):
        order.append("report")
        out = work / "20260819" / "LME每日報價20260819.xlsx"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_bytes(b"out")
        return out

    monkeypatch.setattr("lme_daily.bbg_fetch.fetch_bloomberg_snapshot_and_3m", fake_fetch)
    monkeypatch.setattr("lme_daily.vba_runner.run_reference_macro", fake_vba)
    monkeypatch.setattr("lme_daily.report_builder.build_report", fake_report)

    dest = run(
        config,
        as_of=date(2026, 8, 19),
        dry_run=False,
        skip_vba=False,
        skip_bbg=False,
    )
    assert dest is not None
    assert order[:2] == ["bbg", "vba"]
    assert captured["three_m"] == "2026/12/01"
    assert captured["prev"] == "2026/08/18"
