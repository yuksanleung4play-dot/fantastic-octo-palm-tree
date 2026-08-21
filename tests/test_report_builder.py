"""報告產生（matplotlib / xlsxwriter），不需 Excel COM。"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from lme_daily.config import load_config
from lme_daily.exceptions import ReportBuildError
from lme_daily.report_builder import (
    AUTOFIT_MIN_WIDTH,
    DATE_COL,
    DISCLAIMER_FONT_SIZE,
    HF_FONT_SIZE,
    NUMBER_FORMAT_2DP,
    PRINT_DISCLAIMER_FULL,
    SHEET_BBG,
    SHEET_CHART,
    SHEET_PRINT,
    SHEET_RAW,
    SIX_SERIES,
    SOURCE_LABEL,
    apply_autofit_columns,
    build_report,
    excel_absolute_external_formula,
    header_footer_run,
    load_forward_curve,
    slice_plot_window,
)


def _write_min_config(tmp_path: Path, engine: str) -> Path:
    import yaml

    work = tmp_path / "work"
    work.mkdir(exist_ok=True)
    payload = {
        "paths": {
            "working_dir": str(work),
            "ref_workbook_name": "ref.xlsm",
            "bbg_workbook_name": "bbg.xlsx",
            "output_prefix": "LME每日報價",
            "output_dir": "",
        },
        "vba": {"macro_name": "RunDailyLME", "use_param_injection": True},
        "bloomberg": {
            "copy_range": "B3:I10",
            "bbg_sheet_name": "Promt date",
            "refresh_wait_seconds": 1,
        },
        "chart": {"forward_months": 27, "engine": engine},
        "logging": {"level": "INFO", "file": ""},
    }
    path = tmp_path / "config.yaml"
    path.write_text(yaml.safe_dump(payload, allow_unicode=True), encoding="utf-8")
    return path


def _winish(value: object) -> str:
    return str(value).replace("/", "\\")


def _make_step2(path: Path) -> None:
    cash = datetime(2026, 8, 19)
    rows = []
    # 混雜日期字串 + datetime；NI/SN 在較長天期留空
    for i in range(0, 40):
        d = cash + pd.DateOffset(months=i) if i % 3 == 0 else cash + pd.Timedelta(days=i * 20)
        prompt: object
        if i % 5 == 0:
            prompt = d.strftime("%d-%b-%y")
        elif i % 5 == 1:
            prompt = d.strftime("%d/%b/%y")
        else:
            prompt = d
        row = {
            DATE_COL: prompt,
            "CA": 9000 + i,
            "AH": 2500 + i,
            "ZS": 2800 + i,
            "NI": 16000 + i if i < 20 else None,
            "SN": 30000 + i if i < 18 else None,
            "PB": 2000 + i,
        }
        rows.append(row)
    # 一列無法解析的日期：應記警告但不中斷
    rows.append(
        {
            DATE_COL: "not-a-date",
            "CA": 1,
            "AH": 1,
            "ZS": 1,
            "NI": 1,
            "SN": 1,
            "PB": 1,
        }
    )
    pd.DataFrame(rows).to_excel(path, index=False)


def _bbg_sample():
    values = (
        ("Metal", "Last", "Bid", "Ask", "High", "Low", "Settle", "AsOf"),
        ("CA", 9001.5, 9000.0, 9002.0, 9050.0, 8980.0, 9001.0, 45953),
    )
    formats = (
        ("General",) * 8,
        ("General", "0.0", "0.0", "0.0", "0.0", "0.0", "0.0", "YYYY-MM-DD"),
    )
    return values, formats


@pytest.mark.parametrize("engine", ["matplotlib", "xlsxwriter"])
def test_build_report_three_sheets(tmp_path: Path, engine: str):
    cfg_path = _write_min_config(tmp_path, engine)
    config = load_config(cfg_path)
    as_of = date(2026, 8, 19)
    vba_dir = config.vba_dir(as_of)
    vba_dir.mkdir(parents=True, exist_ok=True)
    step2 = vba_dir / "20260819.xlsx"
    _make_step2(step2)
    values, formats = _bbg_sample()

    dest = build_report(
        config,
        as_of=as_of,
        step2_path=step2,
        bbg_values=values,
        bbg_formats=formats,
    )
    assert dest.name == "LME每日報價20260819.xlsx"
    assert dest.parent == config.run_dir(as_of)
    assert dest.is_file()

    wb = load_workbook(dest)
    assert SHEET_BBG in wb.sheetnames
    assert SHEET_CHART in wb.sheetnames
    assert SHEET_RAW in wb.sheetnames
    assert SHEET_PRINT in wb.sheetnames
    assert wb[SHEET_BBG]["A1"].value == "Metal"
    assert wb[SHEET_BBG]["B2"].value == 9001.5
    # 原始數據應含未過濾的長天期列（含表頭 > 27 列），且為貼值而非外部連結
    raw_rows = wb[SHEET_RAW].max_row
    assert raw_rows >= 40
    raw_a1 = wb[SHEET_RAW]["A1"].value
    assert raw_a1 == DATE_COL
    assert not str(raw_a1).startswith("=")
    assert wb[SHEET_RAW]["B2"].value == 9000
    for row in wb[SHEET_RAW].iter_rows(max_row=raw_rows, max_col=8):
        for cell in row:
            text = str(cell.value or "")
            assert not text.startswith("="), cell.coordinate
            assert ".xlsx" not in text.lower()
    print_ws = wb[SHEET_PRINT]
    assert "LME每日報價" in str(print_ws["A2"].value)
    assert "LME Daily Quotation" in str(print_ws["A3"].value)
    assert "第 &P 頁，共 &N 頁" in (print_ws.oddFooter.center.text or "")
    assert "免責聲明" not in (print_ws.oddFooter.left.text or "")
    assert print_ws.page_setup.orientation == "landscape"
    if engine == "matplotlib":
        assert print_ws.page_setup.fitToWidth == 1
        assert len(wb[SHEET_CHART]._images) == 6
    wb.close()


def test_plot_window_truncates_long_tenor(tmp_path: Path):
    step2 = tmp_path / "curve.xlsx"
    _make_step2(step2)
    df = load_forward_curve(step2)
    plot_df = slice_plot_window(df, forward_months=27)
    cash = plot_df[DATE_COL].min()
    cutoff = cash + pd.DateOffset(months=27)
    assert plot_df[DATE_COL].max() <= cutoff
    assert len(plot_df) < len(df.dropna(subset=[DATE_COL]))
    # 各品種 dropna 不可把空值填 0
    ni = plot_df.dropna(subset=["NI"])
    assert ni["NI"].isna().sum() == 0
    assert (ni["NI"] == 0).sum() == 0
    assert set(SIX_SERIES) <= set(df.columns)


def test_excel_absolute_external_formula_uses_folder_and_file(tmp_path: Path):
    source = tmp_path / "work" / "20260820" / "20260820.xlsx"
    formula = excel_absolute_external_formula(source, "Sheet1", "A1")
    assert formula.startswith("='")
    assert "[20260820.xlsx]Sheet1" in formula
    assert formula.endswith("!A1")
    assert _winish(source.parent) in _winish(formula)


def test_missing_step2_raises(tmp_path: Path):
    cfg_path = _write_min_config(tmp_path, "matplotlib")
    config = load_config(cfg_path)
    as_of = date(2026, 8, 19)
    with pytest.raises(ReportBuildError, match="步驟二產出檔不存在"):
        build_report(
            config,
            as_of=as_of,
            step2_path=config.step2_workbook_path(as_of),
            bbg_values=(),
            bbg_formats=(),
        )


def test_output_dir_does_not_move_vba_or_chart_source(tmp_path: Path):
    import yaml

    work = tmp_path / "work"
    export = tmp_path / "LME_Export"
    work.mkdir()
    payload = {
        "paths": {
            "working_dir": str(work),
            "ref_workbook_name": "ref.xlsm",
            "bbg_workbook_name": "bbg.xlsx",
            "output_prefix": "LME每日報價",
            "output_dir": str(export),
        },
        "vba": {"macro_name": "RunDailyLME", "use_param_injection": True},
        "bloomberg": {
            "copy_range": "B3:I10",
            "bbg_sheet_name": "Promt date",
            "refresh_wait_seconds": 1,
        },
        "chart": {"forward_months": 27, "engine": "matplotlib"},
        "logging": {"level": "INFO", "file": ""},
    }
    cfg = tmp_path / "config.yaml"
    cfg.write_text(yaml.safe_dump(payload, allow_unicode=True), encoding="utf-8")
    config = load_config(cfg)
    as_of = date(2026, 8, 20)
    vba_dir, run_dir = config.ensure_run_dirs(as_of)
    assert vba_dir == work.resolve() / "20260820"
    assert run_dir == export.resolve() / "20260820"
    step2 = vba_dir / "20260820.xlsx"
    _make_step2(step2)
    dest = build_report(
        config,
        as_of=as_of,
        step2_path=step2,
        bbg_values=_bbg_sample()[0],
        bbg_formats=_bbg_sample()[1],
    )
    assert dest.parent == run_dir
    assert dest.is_file()
    assert step2.is_file()
    assert not (run_dir / "20260820.xlsx").exists()

    wb = load_workbook(dest)
    raw_a1 = str(wb[SHEET_RAW]["A1"].value or "")
    chart_src = str(wb[SHEET_CHART]["A2"].value or "")
    wb.close()
    assert raw_a1 == DATE_COL
    assert not raw_a1.startswith("=")
    assert SOURCE_LABEL not in chart_src
    assert "20260820.xlsx" not in chart_src
    assert _winish(vba_dir) not in chart_src
    assert _winish(run_dir) not in chart_src


def _header_footer_parts(ws: Worksheet) -> list[object]:
    parts: list[object] = []
    for hf_name in (
        "oddHeader",
        "evenHeader",
        "firstHeader",
        "oddFooter",
        "evenFooter",
        "firstFooter",
    ):
        hf = getattr(ws, hf_name, None)
        if hf is None:
            continue
        for side in ("left", "center", "right"):
            item = getattr(hf, side, None)
            if item is not None and (getattr(item, "text", None) or getattr(item, "size", None)):
                parts.append(item)
    return parts


def _header_footer_texts(ws: Worksheet) -> list[str]:
    return [str(part.text) for part in _header_footer_parts(ws) if getattr(part, "text", None)]


def _cell_texts(ws: Worksheet, *, max_row: int | None = None, max_col: int = 20) -> list[str]:
    texts: list[str] = []
    for row in ws.iter_rows(max_row=max_row or ws.max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                texts.append(str(cell.value))
    return texts


def test_header_footer_run_keeps_date_out_of_font_size():
    """問題三：``&8`` + 日期不可被 Excel 解析成 ``&82``（82pt）。"""
    blob = header_footer_run("2B2B2B", "2026-08-20")
    assert blob == f"&K2B2B2B&{HF_FONT_SIZE} 2026-08-20"
    assert "&82" not in blob
    assert re.search(r"&8\d", blob) is None
    assert f"&{HF_FONT_SIZE} " in blob


@pytest.mark.parametrize("engine", ["matplotlib", "xlsxwriter"])
def test_chart_sheet_has_no_source_path(tmp_path: Path, engine: str):
    """問題一：遠期走勢圖 sheet 不寫入資料來源路徑。"""
    cfg_path = _write_min_config(tmp_path, engine)
    config = load_config(cfg_path)
    as_of = date(2026, 8, 19)
    vba_dir = config.vba_dir(as_of)
    vba_dir.mkdir(parents=True, exist_ok=True)
    step2 = vba_dir / "20260819.xlsx"
    _make_step2(step2)
    dest = build_report(
        config,
        as_of=as_of,
        step2_path=step2,
        bbg_values=_bbg_sample()[0],
        bbg_formats=_bbg_sample()[1],
    )
    wb = load_workbook(dest)
    chart = wb[SHEET_CHART]
    blobs = _cell_texts(chart) + _header_footer_texts(chart)
    wb.close()
    joined = "\n".join(blobs)
    assert SOURCE_LABEL not in joined
    assert "資料來源" not in joined
    assert "20260819.xlsx" not in joined
    assert _winish(vba_dir) not in _winish(joined)
    assert ".xlsx]" not in joined


@pytest.mark.parametrize("engine", ["matplotlib", "xlsxwriter"])
def test_print_sheet_header_font_size_and_layout(tmp_path: Path, engine: str):
    """問題三：合併版面頁首不可出現巨大字級；PrintArea / 橫向 / 縮放正常。"""
    cfg_path = _write_min_config(tmp_path, engine)
    config = load_config(cfg_path)
    as_of = date(2026, 8, 20)
    vba_dir = config.vba_dir(as_of)
    vba_dir.mkdir(parents=True, exist_ok=True)
    step2 = vba_dir / "20260820.xlsx"
    _make_step2(step2)
    dest = build_report(
        config,
        as_of=as_of,
        step2_path=step2,
        bbg_values=_bbg_sample()[0],
        bbg_formats=_bbg_sample()[1],
    )
    wb = load_workbook(dest)
    ws = wb[SHEET_PRINT]
    hf = "\n".join(_header_footer_texts(ws))
    assert re.search(r"&8\d", hf) is None, hf
    assert "&82" not in hf
    for part in _header_footer_parts(ws):
        size = getattr(part, "size", None)
        if size is not None:
            assert float(size) <= 12, (getattr(part, "text", None), size)
    assert ws.oddHeader.right.text == "2026-08-20"
    assert ws.oddHeader.right.size == 8
    assert ws.print_area
    assert ws.page_setup.orientation == "landscape"
    if engine == "matplotlib":
        assert ws.page_setup.fitToWidth == 1
        assert ws.sheet_properties.pageSetUpPr.fitToPage
        assert not ws._charts
    for row in ws.iter_rows():
        for cell in row:
            size = cell.font.size if cell.font is not None else None
            if size is not None:
                assert float(size) <= 24, (cell.coordinate, size)
    wb.close()


def _find_disclaimer_cell(ws: Worksheet):
    for row in ws.iter_rows():
        for cell in row:
            text = str(cell.value or "")
            if "投資附帶風險" in text and "本報告由山證國際" in text:
                return cell
    return None


@pytest.mark.parametrize("engine", ["matplotlib", "xlsxwriter"])
def test_print_sheet_full_disclaimer_only(tmp_path: Path, engine: str):
    """合併版面底部為完整兩段免責聲明；其他 sheet 不重複。"""
    cfg_path = _write_min_config(tmp_path, engine)
    config = load_config(cfg_path)
    as_of = date(2026, 8, 19)
    vba_dir = config.vba_dir(as_of)
    vba_dir.mkdir(parents=True, exist_ok=True)
    step2 = vba_dir / "20260819.xlsx"
    _make_step2(step2)
    dest = build_report(
        config,
        as_of=as_of,
        step2_path=step2,
        bbg_values=_bbg_sample()[0],
        bbg_formats=_bbg_sample()[1],
    )
    wb = load_workbook(dest)
    print_ws = wb[SHEET_PRINT]
    cell = _find_disclaimer_cell(print_ws)
    assert cell is not None
    text = str(cell.value)
    assert text == PRINT_DISCLAIMER_FULL
    assert "\n\n" in text
    assert text.count("免責聲明") == 1
    assert cell.alignment.wrap_text is True
    size = cell.font.size
    assert size is not None
    assert 7 <= float(size) <= 8
    assert float(size) == float(DISCLAIMER_FONT_SIZE)
    height = print_ws.row_dimensions[cell.row].height
    assert height is not None and float(height) >= 60
    assert str(cell.row) in (print_ws.print_area or "")
    assert "本報告僅供參考" not in text
    for name in (SHEET_BBG, SHEET_CHART, SHEET_RAW):
        ws = wb[name]
        joined = "\n".join(_cell_texts(ws) + _header_footer_texts(ws))
        assert "免責聲明" not in joined
        assert "本報告僅供參考" not in joined
        assert "投資附帶風險" not in joined
    wb.close()


def test_apply_autofit_columns_covers_five_digit_prices():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "CA"
    ws["A2"] = 16554.08
    ws["A2"].number_format = NUMBER_FORMAT_2DP
    ws["B1"] = "NI"
    ws["B2"] = 54454.00
    ws["B2"].number_format = NUMBER_FORMAT_2DP
    apply_autofit_columns(ws)
    assert ws.column_dimensions["A"].width >= AUTOFIT_MIN_WIDTH
    assert ws.column_dimensions["B"].width >= AUTOFIT_MIN_WIDTH
    wb.close()


@pytest.mark.parametrize("engine", ["matplotlib", "xlsxwriter"])
def test_report_truncates_values_and_autofits_columns(tmp_path: Path, engine: str):
    cfg_path = _write_min_config(tmp_path, engine)
    config = load_config(cfg_path)
    as_of = date(2026, 8, 19)
    vba_dir = config.vba_dir(as_of)
    vba_dir.mkdir(parents=True, exist_ok=True)
    step2 = vba_dir / "20260819.xlsx"
    _make_step2(step2)
    src = pd.read_excel(step2)
    for col in ("CA", "NI", "SN"):
        src[col] = src[col].astype(float)
    src.loc[0, "CA"] = 16554.08984375
    src.loc[0, "NI"] = 54454.089
    src.loc[0, "SN"] = 1838.005
    src.to_excel(step2, index=False)

    values = (
        ("Metal", "Last", "Bid", "Ask", "High", "Low", "Settle", "AsOf"),
        ("CA", 16554.08984375, 1838.129999, 1838.005, 9050.0, 8980.0, 9001.0, 45953),
    )
    formats = (
        ("General",) * 8,
        ("General", "0.0", "0.0", "0.0", "0.0", "0.0", "0.0", "YYYY-MM-DD"),
    )
    dest = build_report(
        config,
        as_of=as_of,
        step2_path=step2,
        bbg_values=values,
        bbg_formats=formats,
    )
    wb = load_workbook(dest)
    bbg = wb[SHEET_BBG]
    assert bbg["B2"].value == 16554.08
    assert bbg["C2"].value == 1838.12
    assert bbg["D2"].value == 1838.00
    assert "0.00" in (bbg["B2"].number_format or "")
    assert bbg.column_dimensions["B"].width >= AUTOFIT_MIN_WIDTH

    raw = wb[SHEET_RAW]
    assert raw["B2"].value == 16554.08
    assert raw["E2"].value == 54454.08
    assert raw["F2"].value == 1838.00
    assert "0.00" in (raw["B2"].number_format or "")
    assert raw.column_dimensions["B"].width >= AUTOFIT_MIN_WIDTH
    assert raw.column_dimensions["E"].width >= AUTOFIT_MIN_WIDTH
    assert raw.column_dimensions["F"].width >= AUTOFIT_MIN_WIDTH
    wb.close()


