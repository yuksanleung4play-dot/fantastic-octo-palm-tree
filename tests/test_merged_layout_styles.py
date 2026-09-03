"""合併版面定案樣式：A3 主標題、C8 表頭代碼、A271 免責聲明。"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from lme_daily.config import load_config
from lme_daily.report_builder import (
    AUTOFIT_MIN_WIDTH,
    COLOR_ACCENT,
    COLOR_DATE_BG,
    COLOR_DATE_TEXT,
    COLOR_DISCLAIMER_TEXT,
    COLOR_NAVY,
    COLOR_TEXT_CREAM,
    COLOR_TEXT_WHITE,
    DATE_COL,
    FONT_NAME,
    PRINT_CHART_BUFFER_ROWS,
    PRINT_CHART_ROW_STRIDE,
    PRINT_DISCLAIMER_FULL,
    PRINT_FOOTER,
    PRINT_LAST_COL,
    SHEET_BBG,
    SHEET_PRINT,
    build_report,
    center_image_in_merged_range,
    column_width_to_emu,
    column_width_to_pixels,
    compute_center_anchor_in_range,
    disclaimer_row_after,
    merged_layout_disclaimer_row,
    print_chart_display_pixel_size,
)


def _write_min_config(tmp_path: Path, engine: str, **extra) -> Path:
    import yaml

    work = tmp_path / "work"
    work.mkdir(exist_ok=True)
    payload = {
        "paths": {
            "working_dir": str(work),
            "ref_workbook_name": "ref.xlsm",
            "bbg_workbook_name": "bbg.xlsx",
            "output_prefix": "LME每日報價",
            "output_dir": "",
        },
        "vba": {"macro_name": "RunDailyLME", "use_param_injection": True},
        "bloomberg": {
            "copy_range": "B3:I10",
            "bbg_sheet_name": "Promt date",
            "refresh_wait_seconds": 1,
        },
        "chart": {"forward_months": 27, "engine": engine},
        "logging": {"level": "INFO", "file": ""},
    }
    payload.update(extra)
    path = tmp_path / "config.yaml"
    path.write_text(yaml.safe_dump(payload, allow_unicode=True), encoding="utf-8")
    return path


def _hex_color(color) -> str | None:
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if rgb is None or rgb in ("00000000", "None", "0"):
        return None
    text = str(rgb)
    if text in {"00000000", "None", "0"}:
        return None
    if len(text) >= 6:
        return text[-6:].upper()
    return text.upper()


def fill_hex(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.patternType in (None, "none"):
        return None
    return _hex_color(fill.fgColor)


def font_hex(cell) -> str | None:
    font = cell.font
    if font is None:
        return None
    return _hex_color(font.color)


def _bbg_tenor_sample():
    """對應 B3:I10 的 8×8 表頭代碼列（C:H = CA/AH/PB/NI/SN/ZS）。"""
    values = (
        ("", "", "CA", "AH", "PB", "NI", "SN", "ZS"),
        (
            "Tenors",
            "Prompt date",
            "Copper",
            "Primary Aluminium",
            "Lead",
            "Nickel",
            "Tin",
            "Zinc",
        ),
        ("Cash", date(2026, 9, 1), 14461.57, 3218.10, 1873.41, 16712.76, 54479.00, 4091.95),
        ("3MO", date(2026, 11, 27), 14253.00, 3225.00, 1910.50, 16887.00, 54827.00, 3893.00),
        ("15Mo", date(2027, 11, 17), 14171.68, 3146.00, 2051.03, 17619.50, 55040.00, 3722.13),
        ("27Mo", date(2028, 11, 15), 14138.43, 3072.00, 2144.28, 18241.50, "N/A", 3470.88),
        ("63Mo", date(2031, 11, 19), 14146.43, 2992.00, 2224.78, 20245.50, "N/A", 3154.88),
        ("123Mo", date(2036, 11, 19), 14148.43, 3094.00, "N/A", "N/A", "N/A", "N/A"),
    )
    formats = (("General",) * 8,) * 8
    return values, formats


def _make_step2_rows(path: Path, n_data: int) -> None:
    cash = datetime(2026, 8, 28)
    rows = []
    for i in range(n_data):
        d = cash + pd.Timedelta(days=i)
        rows.append(
            {
                DATE_COL: d,
                "CA": 9000 + i,
                "AH": 2500 + i,
                "ZS": 2800 + i,
                "NI": 16000 + i,
                "SN": 30000 + i,
                "PB": 2000 + i,
            }
        )
    pd.DataFrame(rows).to_excel(path, index=False)


def test_merged_layout_disclaimer_row_matches_fu_ben_a271():
    assert merged_layout_disclaimer_row(bbg_rows=8, raw_rows=210) == 271


@pytest.mark.parametrize("engine", ["matplotlib", "xlsxwriter"])
def test_merged_layout_frozen_styles_a3_c8_a271(tmp_path: Path, engine: str):
    cfg_path = _write_min_config(tmp_path, engine)
    config = load_config(cfg_path)
    as_of = date(2026, 8, 27)
    vba_dir = config.vba_dir(as_of)
    vba_dir.mkdir(parents=True, exist_ok=True)
    step2 = vba_dir / "20260827.xlsx"
    _make_step2_rows(step2, 209)
    values, formats = _bbg_tenor_sample()
    assert merged_layout_disclaimer_row(bbg_rows=len(values), raw_rows=210) == 271

    dest = build_report(
        config,
        as_of=as_of,
        step2_path=step2,
        bbg_values=values,
        bbg_formats=formats,
    )
    wb = load_workbook(dest)
    ws = wb[SHEET_PRINT]

    assert ws["A3"].value == "LME每日報價"
    assert fill_hex(ws["A3"]) == COLOR_ACCENT
    assert font_hex(ws["A3"]) == COLOR_TEXT_CREAM
    assert ws["A3"].font.bold is True
    assert float(ws["A3"].font.size) == 20
    assert ws["A3"].font.name == FONT_NAME
    assert float(ws.row_dimensions[3].height) == 34

    assert ws["C8"].value == "CA"
    assert fill_hex(ws["C8"]) == COLOR_ACCENT
    assert font_hex(ws["C8"]) == COLOR_TEXT_WHITE
    assert ws["C8"].font.bold is True
    assert ws["C8"].alignment.horizontal == "center"
    assert ws["C8"].alignment.vertical == "center"
    assert ws["H8"].value == "ZS"
    assert fill_hex(ws["H8"]) == COLOR_ACCENT
    assert font_hex(ws["H8"]) == COLOR_TEXT_WHITE

    for coord in ("A8", "B8"):
        assert fill_hex(ws[coord]) == COLOR_ACCENT
        assert font_hex(ws[coord]) == COLOR_TEXT_WHITE
        assert ws[coord].font.bold is True
        assert float(ws[coord].font.size) == 11
        assert ws[coord].font.name == FONT_NAME
        assert ws[coord].alignment.horizontal == "center"
        assert ws[coord].alignment.vertical == "center"

    assert ws["A9"].value == "Tenors"
    assert ws["A9"].alignment.horizontal == "center"
    assert ws["A9"].alignment.vertical == "center"
    for label in ("Cash", "3MO", "15Mo", "27Mo", "63Mo", "123Mo"):
        cell = next(
            c
            for row in ws.iter_rows(min_col=1, max_col=1)
            for c in row
            if c.value == label
        )
        assert cell.alignment.horizontal == "center", label
        assert cell.alignment.vertical == "center", label
    assert "Bloomberg Snapshot" in str(ws["A7"].value)
    assert ws["A7"].alignment.horizontal == "left"

    assert fill_hex(ws["A5"]) == COLOR_DATE_BG
    assert font_hex(ws["A5"]) == COLOR_DATE_TEXT

    disc = ws["A271"]
    assert disc.value == PRINT_DISCLAIMER_FULL
    assert fill_hex(disc) is None
    assert font_hex(disc) == COLOR_DISCLAIMER_TEXT
    assert float(disc.font.size) == 8
    assert disc.alignment.wrap_text is True
    assert disc.alignment.horizontal == "left"
    assert disc.alignment.vertical == "top"
    height = ws.row_dimensions[271].height
    assert height is not None and float(height) >= 78

    def _width(letter: str) -> float:
        value = ws.column_dimensions[letter].width
        assert value is not None
        return float(value)

    for letter in "ABCDEFGH":
        assert _width(letter) >= AUTOFIT_MIN_WIDTH, letter

    bbg = wb[SHEET_BBG]
    assert fill_hex(bbg["A1"]) == COLOR_NAVY
    assert font_hex(bbg["A1"]) == COLOR_TEXT_WHITE
    wb.close()


def _write_logo_png(path: Path, *, width: int = 80, height: int = 40) -> Path:
    from PIL import Image

    Image.new("RGB", (width, height), (182, 137, 95)).save(path)
    return path


def _image_in_logo_row(ws) -> bool:
    for img in ws._images:
        anchor = img.anchor
        if isinstance(anchor, str) and str(anchor).upper().startswith("A2"):
            return True
        origin = getattr(anchor, "_from", None)
        if origin is not None and int(origin.row) == 1 and 0 <= int(origin.col) <= 7:
            return True
    return False


def test_logo_pixel_size_keeps_aspect_ratio(tmp_path: Path):
    from lme_daily.report_builder import PRINT_LOGO_ROW_HEIGHT, _logo_pixel_size

    logo = _write_logo_png(tmp_path / "logo.png", width=80, height=40)
    width, height = _logo_pixel_size(logo, row_height_pt=PRINT_LOGO_ROW_HEIGHT)
    assert height == round(PRINT_LOGO_ROW_HEIGHT * 96 / 72)
    assert width == round(80 * height / 40)


def _build_small_report(tmp_path: Path, engine: str, **config_extra):
    cfg_path = _write_min_config(tmp_path, engine, **config_extra)
    config = load_config(cfg_path)
    as_of = date(2026, 8, 27)
    vba_dir = config.vba_dir(as_of)
    vba_dir.mkdir(parents=True, exist_ok=True)
    step2 = vba_dir / "20260827.xlsx"
    _make_step2_rows(step2, 5)
    values, formats = _bbg_tenor_sample()
    dest = build_report(
        config,
        as_of=as_of,
        step2_path=step2,
        bbg_values=values,
        bbg_formats=formats,
    )
    return dest


@pytest.mark.parametrize("engine", ["matplotlib", "xlsxwriter"])
def test_print_sheet_embeds_floating_logo_at_a2(tmp_path: Path, engine: str):
    logo = _write_logo_png(tmp_path / "company-logo.png")
    dest = _build_small_report(tmp_path, engine, branding={"logo_path": str(logo)})
    wb = load_workbook(dest)
    ws = wb[SHEET_PRINT]
    texts = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
    assert all("DISPIMG" not in text and "_xlfn.DISPIMG" not in text for text in texts)
    assert "A2:H2" in {str(rng) for rng in ws.merged_cells.ranges}
    assert float(ws.row_dimensions[2].height) == 34
    assert _image_in_logo_row(ws)
    wb.close()


@pytest.mark.parametrize("engine", ["matplotlib", "xlsxwriter"])
def test_print_sheet_skips_logo_when_unset(tmp_path: Path, engine: str, caplog):
    import logging

    with caplog.at_level(logging.WARNING, logger="lme_daily.report_builder"):
        dest = _build_small_report(tmp_path, engine)
    assert dest.is_file()
    assert any("未設定 Logo 路徑，略過插入" in rec.message for rec in caplog.records)
    wb = load_workbook(dest)
    ws = wb[SHEET_PRINT]
    assert not _image_in_logo_row(ws)
    wb.close()


@pytest.mark.parametrize("engine", ["matplotlib", "xlsxwriter"])
def test_print_sheet_skips_logo_when_file_missing(tmp_path: Path, engine: str, caplog):
    import logging

    missing = tmp_path / "no-such-logo.png"
    with caplog.at_level(logging.WARNING, logger="lme_daily.report_builder"):
        dest = _build_small_report(tmp_path, engine, branding={"logo_path": str(missing)})
    assert dest.is_file()
    assert any("找不到 Logo 檔案，略過插入" in rec.message for rec in caplog.records)
    wb = load_workbook(dest)
    assert not _image_in_logo_row(wb[SHEET_PRINT])
    wb.close()


def test_column_width_to_emu_uses_openpyxl_pixels_to_emu():
    from openpyxl.utils.units import pixels_to_EMU

    assert column_width_to_pixels(10) == 10 * 7 + 5
    assert column_width_to_emu(10) == pixels_to_EMU(10 * 7 + 5)


def test_center_image_in_merged_range_spans_columns_when_offset_exceeds_first(tmp_path: Path):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
    from openpyxl.utils.units import pixels_to_EMU

    logo = _write_logo_png(tmp_path / "logo.png", width=80, height=40)
    wb = Workbook()
    ws = wb.active
    for letter in "ABCDEFGH":
        ws.column_dimensions[letter].width = 10
    ws.row_dimensions[2].height = 34
    ws.merge_cells("A2:H2")
    img = XLImage(str(logo))
    img.width = 90
    img.height = 45
    center_image_in_merged_range(ws, img, "A2:H2")
    assert isinstance(img.anchor, OneCellAnchor)
    col_0, col_off, row_0, row_off = compute_center_anchor_in_range(
        col_widths_chars=[10] * 8,
        row_heights_pt=[34],
        image_width_px=90,
        image_height_px=45,
        start_col_0=0,
        start_row_0=1,
    )
    assert img.anchor._from.col == col_0
    assert int(img.anchor._from.colOff) == col_off
    assert img.anchor._from.row == row_0 == 1
    assert int(img.anchor._from.rowOff) == row_off
    assert col_0 > 0
    assert col_off < column_width_to_emu(10)
    assert img.anchor.ext.cx == pixels_to_EMU(90)
    assert img.anchor.ext.cy == pixels_to_EMU(45)
    assert img in ws._images
    wb.close()


def _print_area_end_row(ws) -> int:
    area = ws.print_area
    assert area, "print_area should be set"
    text = area if isinstance(area, str) else ",".join(str(part) for part in area)
    matches = re.findall(r"\$?[A-Za-z]+\$?(\d+)", text)
    assert matches, text
    return int(matches[-1])


def _last_raw_content_row(ws) -> int:
    from lme_daily.report_builder import PRINT_RAW_SECTION_TITLE

    heading = next(
        cell
        for row in ws.iter_rows()
        for cell in row
        if cell.value == PRINT_RAW_SECTION_TITLE
    )
    last = heading.row
    for row in ws.iter_rows(min_row=heading.row + 1, max_col=PRINT_LAST_COL):
        values = [cell.value for cell in row]
        if any(isinstance(v, str) and "免責聲明" in v for v in values if v is not None):
            break
        if any(v is not None for v in values):
            last = row[0].row
    return last


@pytest.mark.parametrize("engine", ["matplotlib", "xlsxwriter"])
@pytest.mark.parametrize("n_data", [100, 300])
def test_disclaimer_follows_raw_content_length(tmp_path: Path, engine: str, n_data: int):
    cfg_path = _write_min_config(tmp_path, engine)
    config = load_config(cfg_path)
    as_of = date(2026, 8, 27)
    vba_dir = config.vba_dir(as_of)
    vba_dir.mkdir(parents=True, exist_ok=True)
    step2 = vba_dir / "20260827.xlsx"
    _make_step2_rows(step2, n_data)
    values, formats = _bbg_tenor_sample()
    dest = build_report(
        config,
        as_of=as_of,
        step2_path=step2,
        bbg_values=values,
        bbg_formats=formats,
    )
    wb = load_workbook(dest)
    ws = wb[SHEET_PRINT]
    disc = next(
        cell
        for row in ws.iter_rows()
        for cell in row
        if cell.value == PRINT_DISCLAIMER_FULL
    )
    last_raw = _last_raw_content_row(ws)
    assert disc.row == last_raw + 2
    expected = merged_layout_disclaimer_row(bbg_rows=len(values), raw_rows=n_data + 1)
    assert disc.row == expected
    assert disc.row == disclaimer_row_after(last_raw)
    assert disc.row != 271
    assert _print_area_end_row(ws) == disc.row
    assert f"H{disc.row}" in str(ws.print_area).replace("$", "")
    footer = ws.oddFooter.center.text or ""
    assert PRINT_FOOTER in footer or ("&P" in footer and "&N" in footer)
    footer_size = ws.oddFooter.center.size
    if footer_size is not None:
        assert int(footer_size) == 9
    wb.close()


def test_a8_b8_match_c8_accent_on_code_row():
    from openpyxl import Workbook

    from lme_daily.report_builder import _write_print_bbg_openpyxl

    wb = Workbook()
    ws = wb.active
    values, formats = _bbg_tenor_sample()
    _write_print_bbg_openpyxl(ws, values, formats, start_row=8)
    for coord in ("A8", "B8", "C8"):
        assert fill_hex(ws[coord]) == COLOR_ACCENT
        assert font_hex(ws[coord]) == COLOR_TEXT_WHITE
        assert ws[coord].font.bold is True
        assert ws[coord].alignment.horizontal == "center"
        assert ws[coord].alignment.vertical == "center"
    wb.close()


def test_print_chart_height_fits_row_gap_with_buffer():
    from openpyxl.utils.units import DEFAULT_ROW_HEIGHT, points_to_pixels

    _width, height = print_chart_display_pixel_size()
    slot_px = points_to_pixels(DEFAULT_ROW_HEIGHT * PRINT_CHART_ROW_STRIDE)
    usable_px = points_to_pixels(
        DEFAULT_ROW_HEIGHT * (PRINT_CHART_ROW_STRIDE - PRINT_CHART_BUFFER_ROWS)
    )
    assert height <= usable_px
    assert height < slot_px

