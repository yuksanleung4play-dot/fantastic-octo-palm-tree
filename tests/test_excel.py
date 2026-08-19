from datetime import date
from pathlib import Path

from openpyxl import Workbook

from lme_calendar.excel import MANAGED_SHEETS, update_workbook
from lme_calendar.holidays import load_holiday_csv

HOLIDAYS_PATH = Path(__file__).resolve().parents[1] / "data" / "holidays.csv"


def test_update_does_not_touch_unmanaged_sheets(tmp_path):
    calendar = load_holiday_csv(HOLIDAYS_PATH)
    path = tmp_path / "book.xlsx"

    wb = Workbook()
    wb.active.title = "Bloomberg_BDP"
    ws = wb["Bloomberg_BDP"]
    ws["A1"] = '=BDP("CAHDS03 Comdty","PX_LAST")'
    ws["B2"] = "do-not-touch"
    wb.save(path)

    result = update_workbook(
        path,
        calendar=calendar,
        as_of=date(2026, 8, 19),
        months_forward=27,
        snapshot_mode="overwrite",
    )

    from openpyxl import load_workbook

    out = load_workbook(path)
    assert "Bloomberg_BDP" in out.sheetnames
    assert out["Bloomberg_BDP"]["A1"].value == '=BDP("CAHDS03 Comdty","PX_LAST")'
    assert out["Bloomberg_BDP"]["B2"].value == "do-not-touch"
    for name in MANAGED_SHEETS:
        assert name in out.sheetnames
    assert out["Cash_3M_Daily"]["A1"].value == "Snapshot_Date"
    assert out["Cash_3M_Daily"]["A2"].value.date() == date(2026, 8, 19)
    assert result["cash_date"] == date(2026, 8, 21)


def test_append_mode_accumulates_and_replaces_same_day(tmp_path):
    calendar = load_holiday_csv(HOLIDAYS_PATH)
    path = tmp_path / "book.xlsx"

    update_workbook(
        path, calendar=calendar, as_of=date(2026, 8, 18), snapshot_mode="append"
    )
    update_workbook(
        path, calendar=calendar, as_of=date(2026, 8, 19), snapshot_mode="append"
    )
    update_workbook(
        path, calendar=calendar, as_of=date(2026, 8, 19), snapshot_mode="append"
    )

    from openpyxl import load_workbook

    out = load_workbook(path)
    ws = out["Cash_3M_Daily"]
    assert ws.max_row == 3  # header + two snapshot dates
    assert ws["A2"].value.date() == date(2026, 8, 18)
    assert ws["A3"].value.date() == date(2026, 8, 19)
