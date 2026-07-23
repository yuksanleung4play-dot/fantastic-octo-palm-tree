#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Create a cross-year reusable department shift roster Excel template (v2)."""

from datetime import date

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DISPLAY_YEAR = 2026
BASE_DATE = date(2026, 6, 29)  # Permanent rotation anchor (Monday)
DEFAULT_AL_QUOTA = 12
NUM_WEEKS = 53
LEAVE_ROWS = 500  # capacity for 請假登記 list
HOLIDAY_ROWS = 80

EMPLOYEES = [
    # name, role (S=fixed, R=rotate, V=vacant), offset, al_quota
    ("ARTHUR", "S", None, DEFAULT_AL_QUOTA),
    ("PATRICK", "R", 0, DEFAULT_AL_QUOTA),
    ("CARREY", "R", 4, DEFAULT_AL_QUOTA),
    ("CESC", "R", 8, DEFAULT_AL_QUOTA),
    ("LUCAS", "R", 12, DEFAULT_AL_QUOTA),
    ("", "V", 0, None),  # vacant
]

HOLIDAYS_2026 = [
    (date(2026, 1, 1), "元旦"),
    (date(2026, 2, 17), "農曆年初一"),
    (date(2026, 2, 18), "農曆年初二"),
    (date(2026, 2, 19), "農曆年初三"),
    (date(2026, 4, 3), "耶穌受難節前一日"),
    (date(2026, 4, 4), "耶穌受難節"),
    (date(2026, 4, 6), "復活節星期一"),
    (date(2026, 5, 1), "勞動節"),
    (date(2026, 5, 25), "佛誕"),
    (date(2026, 6, 19), "端午節"),
    (date(2026, 7, 1), "香港特別行政區成立紀念日"),
    (date(2026, 9, 26), "中秋節翌日"),
    (date(2026, 10, 1), "國慶日"),
    (date(2026, 10, 19), "重陽節"),
    (date(2026, 12, 25), "聖誕節"),
    (date(2026, 12, 26), "聖誕節後第一個周日"),
]

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_SUBHDR = PatternFill("solid", fgColor="2E75B6")
FILL_SETTINGS = PatternFill("solid", fgColor="D6EAF8")
FILL_QUICK = PatternFill("solid", fgColor="FFF2CC")
FILL_QUICK_HDR = PatternFill("solid", fgColor="F4B183")
FILL_S = PatternFill("solid", fgColor="FFFFFF")
FILL_A = PatternFill("solid", fgColor="D9D9D9")
FILL_P = PatternFill("solid", fgColor="C6EFCE")
FILL_N = PatternFill("solid", fgColor="BDD7EE")
FILL_AL = PatternFill("solid", fgColor="FFFF00")
FILL_SL = PatternFill("solid", fgColor="F4B183")
FILL_HOLIDAY = PatternFill("solid", fgColor="E2D5F1")
FILL_VACANT = PatternFill("solid", fgColor="F2F2F2")
FILL_STAT = PatternFill("solid", fgColor="E2EFDA")
FILL_WARN = PatternFill("solid", fgColor="FF6B6B")
FILL_INPUT = PatternFill("solid", fgColor="FFF9E6")
FILL_ANCHOR = PatternFill("solid", fgColor="FCE4D6")
FILL_READONLY = PatternFill("solid", fgColor="F8F9FA")

YEAR_CELL = "'設定'!$B$3"
BASE_CELL = "'設定'!$B$4"
EMP_FIRST_ROW = 8  # 設定!B8:B13 names

FONT_WHITE = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)
FONT_TITLE = Font(name="Microsoft JhengHei", bold=True, size=16, color="1F4E79")
FONT_HDR = Font(name="Microsoft JhengHei", bold=True, size=10, color="FFFFFF")
FONT_NORMAL = Font(name="Microsoft JhengHei", size=10)
FONT_BOLD = Font(name="Microsoft JhengHei", bold=True, size=10)
FONT_SMALL = Font(name="Microsoft JhengHei", size=8)
FONT_NOTE = Font(name="Microsoft JhengHei", size=9, italic=True, color="C00000")
FONT_TINY = Font(name="Microsoft JhengHei", size=7)

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
HOLIDAY_BORDER = Border(
    left=Side(style="medium", color="8E7CC3"),
    right=Side(style="medium", color="8E7CC3"),
    top=Side(style="medium", color="8E7CC3"),
    bottom=Side(style="medium", color="8E7CC3"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Roster layout: A name, B AL used, C AL left, D comp bal, E open, F year new, G+ weeks
FIRST_WEEK_COL = 7  # column G


def week_start_formula(week_index: int) -> str:
    first = (
        f"DATE({YEAR_CELL},1,1)"
        f"+MOD(8-WEEKDAY(DATE({YEAR_CELL},1,1),2),7)"
    )
    week = f"({first})+{week_index}*7"
    return f'=IF(({week})>DATE({YEAR_CELL},12,31),"",{week})'


def rotate_shift_expr(week_monday_ref: str, offset_cell: str) -> str:
    """A/P/N from 16-week cycle given a Monday date ref and offset cell."""
    mod = (
        f"MOD(INT(({week_monday_ref}-{BASE_CELL})/7)"
        f"+IF({offset_cell}=\"\",0,{offset_cell}),16)"
    )
    return (
        f'IF({mod}<4,"A",IF({mod}<8,"P",IF({mod}<12,"N","P")))'
    )


def planned_shift_formula(date_ref: str, set_row: int) -> str:
    """Planned S/A/P/N for a calendar date (uses that date's Monday)."""
    name_c = f"'設定'!$B${set_row}"
    role_c = f"'設定'!$C${set_row}"
    off_c = f"'設定'!$D${set_row}"
    monday = f"({date_ref}-WEEKDAY({date_ref},2)+1)"
    return (
        f'IF({name_c}="","",'
        f'IF({role_c}="S","S",'
        f'IF({role_c}="R",{rotate_shift_expr(monday, off_c)},"")))'
    )


def leave_lookup_formula(name_ref: str, date_ref: str) -> str:
    """Return AL/SL from 請假登記 for name+date, else blank."""
    # SUMPRODUCT / INDEX-MATCH style via COUNTIFS + helper:
    # Prefer looking up 假別 where name and date match.
    return (
        f'IFERROR(INDEX(LeaveType,MATCH(1,('
        f'(LeaveName={name_ref})*(LeaveDate={date_ref})),0)),"")'
    )


def leave_or_shift_day_formula(name_ref: str, date_ref: str, set_row: int) -> str:
    """Display leave if registered, else planned shift. Blank if no date/name."""
    leave = (
        f'IFERROR(INDEX(LeaveType,MATCH(1,('
        f'(LeaveName={name_ref})*(LeaveDate={date_ref})),0)),"")'
    )
    planned = planned_shift_formula(date_ref, set_row)
    return (
        f'IF(OR({date_ref}="",{name_ref}=""),"",'
        f'IF({leave}<>"",{leave},{planned}))'
    )


def add_shift_cf(ws, range_str: str) -> None:
    for code, fill in [
        ("S", FILL_S),
        ("A", FILL_A),
        ("P", FILL_P),
        ("N", FILL_N),
        ("AL", FILL_AL),
        ("SL", FILL_SL),
    ]:
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator="equal", formula=[f'"{code}"'], fill=fill),
        )


def build_workbook():
    wb = Workbook()

    # =====================================================================
    # 設定
    # =====================================================================
    ws_set = wb.active
    ws_set.title = "設定"

    ws_set["A1"] = "部門更表 — 設定區"
    ws_set["A1"].font = FONT_TITLE
    ws_set.merge_cells("A1:G1")

    ws_set["A3"] = "顯示年份"
    ws_set["A3"].font = FONT_BOLD
    ws_set["B3"] = DISPLAY_YEAR
    ws_set["B3"].font = Font(name="Microsoft JhengHei", bold=True, size=14, color="C00000")
    ws_set["B3"].fill = FILL_INPUT
    ws_set["B3"].border = THIN
    ws_set["B3"].alignment = CENTER
    ws_set["C3"] = "← 每年年頭只改此數字，全年週份日期與班次自動重算"
    ws_set["C3"].font = FONT_NOTE

    ws_set["A4"] = "輪班基準日（永久固定）"
    ws_set["A4"].font = FONT_BOLD
    ws_set["B4"] = BASE_DATE
    ws_set["B4"].number_format = "YYYY-MM-DD"
    ws_set["B4"].fill = FILL_ANCHOR
    ws_set["B4"].border = THIN
    ws_set["B4"].font = FONT_BOLD
    ws_set["C4"] = "← 永遠不要修改！所有輪班以此日為錨點跨年連續計算"
    ws_set["C4"].font = FONT_NOTE

    ws_set["A5"] = "說明"
    ws_set["B5"] = (
        "16週循環：4週A → 4週P → 4週N → 4週P。Offset 錯開 0/4/8/12。"
        "年假額度改為每人獨立設定（見下表）。請假請只在「請假登記」輸入。"
    )
    ws_set.merge_cells("B5:G5")
    ws_set["B5"].font = FONT_SMALL

    headers = ["編號", "員工姓名", "班次類型", "初始Offset", "年假額度（天）", "說明"]
    for i, h in enumerate(headers, 1):
        cell = ws_set.cell(7, i, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN

    notes = [
        "固定 S 班，全年不變",
        "輪班 offset=0",
        "輪班 offset=4",
        "輪班 offset=8",
        "輪班 offset=12",
        "填姓名後類型改 R/S，並設定 Offset 與年假額度",
    ]
    for i, (name, role, offset, quota) in enumerate(EMPLOYEES):
        row = EMP_FIRST_ROW + i
        ws_set.cell(row, 1, i + 1).border = THIN
        ws_set.cell(row, 1).alignment = CENTER

        c_name = ws_set.cell(row, 2, name if name else None)
        c_name.fill = FILL_INPUT
        c_name.border = THIN
        c_name.font = FONT_BOLD
        c_name.alignment = CENTER

        c_role = ws_set.cell(row, 3, role)
        c_role.fill = FILL_INPUT
        c_role.border = THIN
        c_role.alignment = CENTER

        c_off = ws_set.cell(row, 4, offset if offset is not None else None)
        c_off.fill = FILL_INPUT
        c_off.border = THIN
        c_off.alignment = CENTER

        c_q = ws_set.cell(row, 5, quota)
        c_q.fill = FILL_INPUT
        c_q.border = THIN
        c_q.alignment = CENTER

        c_note = ws_set.cell(row, 6, notes[i])
        c_note.font = FONT_SMALL
        c_note.border = THIN

        if role == "V":
            for c in range(1, 7):
                if c not in (2, 3, 4, 5):
                    ws_set.cell(row, c).fill = FILL_VACANT

    dv_role = DataValidation(type="list", formula1='"S,R,V"', allow_blank=True)
    ws_set.add_data_validation(dv_role)
    dv_role.add(f"C{EMP_FIRST_ROW}:C{EMP_FIRST_ROW + 5}")

    # Legend
    ws_set["A15"] = "班次／假別顏色圖例（與排班表、月曆一致）"
    ws_set["A15"].font = FONT_BOLD
    legends = [
        ("S", "正常班", FILL_S),
        ("A", "早班", FILL_A),
        ("P", "中班", FILL_P),
        ("N", "晚班", FILL_N),
        ("AL", "年假", FILL_AL),
        ("SL", "病假", FILL_SL),
    ]
    for i, (code, desc, fill) in enumerate(legends):
        ws_set.cell(16, 1 + i, code).fill = fill
        ws_set.cell(16, 1 + i).border = THIN
        ws_set.cell(16, 1 + i).alignment = CENTER
        ws_set.cell(16, 1 + i).font = FONT_BOLD
        ws_set.cell(17, 1 + i, desc).font = FONT_SMALL
        ws_set.cell(17, 1 + i).alignment = CENTER

    ws_set["A19"] = "輪班公式"
    ws_set["A19"].font = FONT_BOLD
    ws_set["A20"] = (
        "週序=INT((該週一−基準日)/7)+Offset；MOD(週序,16)：0–3→A，4–7→P，8–11→N，12–15→P"
    )
    ws_set.merge_cells("A20:G20")
    ws_set["A20"].font = FONT_SMALL

    for col, w in zip("ABCDEFG", [10, 14, 10, 12, 14, 50, 12]):
        ws_set.column_dimensions[col].width = w

    # Named ranges for employee names (for dropdowns)
    wb.defined_names.add(
        DefinedName(name="EmpNames", attr_text=f"'設定'!$B${EMP_FIRST_ROW}:$B${EMP_FIRST_ROW + 5}")
    )

    # =====================================================================
    # 公眾假期
    # =====================================================================
    ws_hol = wb.create_sheet("公眾假期")
    ws_hol["A1"] = "公眾假期資料表"
    ws_hol["A1"].font = FONT_TITLE
    ws_hol.merge_cells("A1:C1")
    ws_hol["A2"] = "每年年頭在此表底部追加新一年假期即可。請勿刪除表頭。"
    ws_hol["A2"].font = FONT_NOTE
    ws_hol.merge_cells("A2:C2")

    for i, h in enumerate(["日期", "假期名稱", "年份（自動）"], 1):
        cell = ws_hol.cell(4, i, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN

    for i, (d, name) in enumerate(HOLIDAYS_2026):
        row = 5 + i
        ws_hol.cell(row, 1, d).number_format = "YYYY-MM-DD"
        ws_hol.cell(row, 1).border = THIN
        ws_hol.cell(row, 1).fill = FILL_HOLIDAY
        ws_hol.cell(row, 2, name).border = THIN
        ws_hol.cell(row, 3, f'=IF(A{row}="","",YEAR(A{row}))').border = THIN

    for row in range(5 + len(HOLIDAYS_2026), 5 + HOLIDAY_ROWS):
        ws_hol.cell(row, 1).number_format = "YYYY-MM-DD"
        ws_hol.cell(row, 1).border = THIN
        ws_hol.cell(row, 1).fill = FILL_INPUT
        ws_hol.cell(row, 2).border = THIN
        ws_hol.cell(row, 2).fill = FILL_INPUT
        ws_hol.cell(row, 3, f'=IF(A{row}="","",YEAR(A{row}))').border = THIN

    ws_hol.column_dimensions["A"].width = 14
    ws_hol.column_dimensions["B"].width = 36
    ws_hol.column_dimensions["C"].width = 14

    wb.defined_names.add(DefinedName(name="HolidayDates", attr_text="'公眾假期'!$A$5:$A$84"))
    wb.defined_names.add(DefinedName(name="HolidayNames", attr_text="'公眾假期'!$B$5:$B$84"))

    # =====================================================================
    # 請假登記（唯一請假輸入源）
    # =====================================================================
    ws_leave = wb.create_sheet("請假登記")
    ws_leave["A1"] = "請假登記（全年唯一請假輸入源）"
    ws_leave["A1"].font = FONT_TITLE
    ws_leave.merge_cells("A1:D1")
    ws_leave["A2"] = (
        "每一行＝一位員工某一天的請假。連續多日請逐日新增多行。"
        "排班表／月曆的 AL・SL 顯示與年假統計皆從此表用公式查找，請勿在其他表輸入請假。"
    )
    ws_leave["A2"].font = FONT_NOTE
    ws_leave.merge_cells("A2:D2")

    for i, h in enumerate(["員工姓名", "請假日期", "假別"], 1):
        cell = ws_leave.cell(4, i, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN

    # Pre-format data rows (table needs at least one data row)
    for row in range(5, 5 + LEAVE_ROWS):
        for col in range(1, 4):
            cell = ws_leave.cell(row, col)
            cell.border = THIN
            cell.fill = FILL_INPUT
            cell.alignment = CENTER
        ws_leave.cell(row, 2).number_format = "YYYY-MM-DD"

    # Sample empty first row kept for table; optional demo left blank

    table = Table(displayName="LeaveTable", ref=f"A4:C{4 + LEAVE_ROWS}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws_leave.add_table(table)

    wb.defined_names.add(DefinedName(name="LeaveName", attr_text="'請假登記'!$A$5:$A$504"))
    wb.defined_names.add(DefinedName(name="LeaveDate", attr_text="'請假登記'!$B$5:$B$504"))
    wb.defined_names.add(DefinedName(name="LeaveType", attr_text="'請假登記'!$C$5:$C$504"))

    dv_emp = DataValidation(
        type="list",
        formula1=f"=EmpNames",
        allow_blank=True,
        showDropDown=False,
    )
    dv_emp.error = "請選擇設定區中的員工姓名"
    dv_emp.errorTitle = "姓名"
    ws_leave.add_data_validation(dv_emp)
    dv_emp.add(f"A5:A{4 + LEAVE_ROWS}")

    dv_type = DataValidation(type="list", formula1='"AL,SL"', allow_blank=True)
    ws_leave.add_data_validation(dv_type)
    dv_type.add(f"C5:C{4 + LEAVE_ROWS}")

    ws_leave.column_dimensions["A"].width = 14
    ws_leave.column_dimensions["B"].width = 14
    ws_leave.column_dimensions["C"].width = 10
    ws_leave.column_dimensions["D"].width = 40
    ws_leave.freeze_panes = "A5"

    ws_leave["E4"] = "提示"
    ws_leave["E4"].font = FONT_BOLD
    ws_leave["E5"] = "假別只能填 AL 或 SL；日期請填實際請假當天（非週一）。"
    ws_leave["E5"].font = FONT_SMALL

    # =====================================================================
    # 假期上班登記
    # =====================================================================
    ws_hw = wb.create_sheet("假期上班登記")
    ws_hw["A1"] = "公眾假期上班登記"
    ws_hw["A1"].font = FONT_TITLE
    ws_hw.merge_cells("A1:H1")
    ws_hw["A2"] = (
        "日期／假期名稱自動由「公眾假期」帶入。請在「上班人員1～6」以手動（下拉）登記當日實際上班者。"
        "補假次數由此表統計，排班表補假欄會自動引用。"
    )
    ws_hw["A2"].font = FONT_NOTE
    ws_hw.merge_cells("A2:H2")

    hw_headers = ["日期", "假期名稱", "上班人員1", "上班人員2", "上班人員3", "上班人員4", "上班人員5", "上班人員6"]
    for i, h in enumerate(hw_headers, 1):
        cell = ws_hw.cell(4, i, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN

    # Pull holidays with formulas (up to HOLIDAY_ROWS)
    for i in range(HOLIDAY_ROWS):
        row = 5 + i
        src = 5 + i
        ws_hw.cell(row, 1, f"=IF('公眾假期'!A{src}=\"\",\"\",'公眾假期'!A{src})")
        ws_hw.cell(row, 1).number_format = "YYYY-MM-DD"
        ws_hw.cell(row, 1).border = THIN
        ws_hw.cell(row, 1).fill = FILL_READONLY
        ws_hw.cell(row, 2, f"=IF('公眾假期'!B{src}=\"\",\"\",'公眾假期'!B{src})")
        ws_hw.cell(row, 2).border = THIN
        ws_hw.cell(row, 2).fill = FILL_READONLY
        for col in range(3, 9):
            cell = ws_hw.cell(row, col)
            cell.border = THIN
            cell.fill = FILL_INPUT
            cell.alignment = CENTER

    dv_hw = DataValidation(type="list", formula1="=EmpNames", allow_blank=True)
    ws_hw.add_data_validation(dv_hw)
    dv_hw.add(f"C5:H{4 + HOLIDAY_ROWS}")

    # 補假統計區
    stat_start = 5 + HOLIDAY_ROWS + 2  # row 87
    ws_hw.cell(stat_start, 1, "補假統計區（自動計算）")
    ws_hw.cell(stat_start, 1).font = FONT_BOLD
    ws_hw.cell(stat_start, 1).fill = FILL_SUBHDR
    ws_hw.cell(stat_start, 1).font = FONT_WHITE
    ws_hw.merge_cells(start_row=stat_start, start_column=1, end_row=stat_start, end_column=4)

    for i, h in enumerate(["員工姓名", "公眾假期上班次數（全部）", "本年上班次數", "說明"], 1):
        cell = ws_hw.cell(stat_start + 1, i, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.border = THIN
        cell.alignment = CENTER

    # Work name range for counting
    work_range = f"$C$5:$H${4 + HOLIDAY_ROWS}"
    date_range = f"$A$5:$A${4 + HOLIDAY_ROWS}"

    for i in range(6):
        row = stat_start + 2 + i
        set_row = EMP_FIRST_ROW + i
        ws_hw.cell(row, 1, f"='設定'!B{set_row}")
        ws_hw.cell(row, 1).border = THIN
        ws_hw.cell(row, 1).font = FONT_BOLD
        # Total count across all registered holiday work columns
        ws_hw.cell(
            row,
            2,
            f'=IF(A{row}="","",COUNTIF({work_range},A{row}))',
        )
        ws_hw.cell(row, 2).border = THIN
        ws_hw.cell(row, 2).fill = FILL_STAT
        ws_hw.cell(row, 2).alignment = CENTER
        # This year only
        ws_hw.cell(
            row,
            3,
            f'=IF(A{row}="","",'
            f'SUMPRODUCT(({date_range}<>"")*(YEAR({date_range})={YEAR_CELL})*'
            f'(($C$5:$C${4 + HOLIDAY_ROWS}=A{row})+($D$5:$D${4 + HOLIDAY_ROWS}=A{row})+'
            f'($E$5:$E${4 + HOLIDAY_ROWS}=A{row})+($F$5:$F${4 + HOLIDAY_ROWS}=A{row})+'
            f'($G$5:$G${4 + HOLIDAY_ROWS}=A{row})+($H$5:$H${4 + HOLIDAY_ROWS}=A{row}))))',
        )
        ws_hw.cell(row, 3).border = THIN
        ws_hw.cell(row, 3).fill = FILL_STAT
        ws_hw.cell(row, 3).alignment = CENTER
        ws_hw.cell(row, 4, "排班表「本年新增補假」引用本欄；累積＝期初＋本年").font = FONT_SMALL
        ws_hw.cell(row, 4).border = THIN

    # Named ranges for roster to pull year counts
    # Emp1 year count at stat_start+2, col 3
    for i in range(6):
        wb.defined_names.add(
            DefinedName(
                name=f"CompYear{i + 1}",
                attr_text=f"'假期上班登記'!$C${stat_start + 2 + i}",
            )
        )

    for col, w in zip("ABCDEFGH", [12, 28, 12, 12, 12, 12, 12, 12]):
        ws_hw.column_dimensions[col].width = w

    ws_hw.freeze_panes = "A5"

    # =====================================================================
    # 排班表
    # =====================================================================
    ws = wb.create_sheet("排班表", 0)

    ws["A1"] = "部門輪班更表（跨年模版）"
    ws["A1"].font = FONT_TITLE
    ws.merge_cells("A1:E1")

    ws["F1"] = "顯示年份"
    ws["F1"].font = FONT_BOLD
    ws["G1"] = f"={YEAR_CELL}"
    ws["G1"].font = Font(name="Microsoft JhengHei", bold=True, size=14, color="C00000")
    ws["G1"].fill = FILL_QUICK
    ws["G1"].border = THIN
    ws["G1"].alignment = CENTER

    # Quick view
    ws["A3"] = "本週／下週快覽"
    ws["A3"].font = FONT_BOLD
    ws["A3"].fill = FILL_QUICK_HDR
    ws.merge_cells("A3:C3")

    ws["A4"] = "今天"
    ws["B4"] = "=TODAY()"
    ws["B4"].number_format = "YYYY-MM-DD"
    ws["B4"].font = FONT_BOLD

    ws["A5"] = "本週起始（一）"
    ws["B5"] = "=B4-WEEKDAY(B4,2)+1"
    ws["B5"].number_format = "YYYY-MM-DD"
    ws["B5"].fill = FILL_QUICK
    ws["B5"].border = THIN

    ws["A6"] = "下週起始（一）"
    ws["B6"] = "=B5+7"
    ws["B6"].number_format = "YYYY-MM-DD"
    ws["B6"].fill = FILL_QUICK
    ws["B6"].border = THIN

    header_row = 13
    name_row0 = 14
    first_col_letter = get_column_letter(FIRST_WEEK_COL)
    last_col_letter = get_column_letter(FIRST_WEEK_COL + NUM_WEEKS - 1)
    week_header_range = f"${first_col_letter}${header_row}:${last_col_letter}${header_row}"

    ws["D5"] = "本週欄位"
    ws["E5"] = f'=IFERROR(MATCH(B5,{week_header_range},0),"非本年")'
    ws["E5"].fill = FILL_QUICK
    ws["E5"].border = THIN

    ws["D6"] = "下週欄位"
    ws["E6"] = f'=IFERROR(MATCH(B6,{week_header_range},0),"非本年")'
    ws["E6"].fill = FILL_QUICK
    ws["E6"].border = THIN

    ws["A8"] = "人員"
    ws["B8"] = "本週班次"
    ws["C8"] = "下週班次"
    for col in range(1, 4):
        cell = ws.cell(8, col)
        cell.fill = FILL_QUICK_HDR
        cell.font = FONT_BOLD
        cell.border = THIN
        cell.alignment = CENTER

    for i in range(6):
        row = 9 + i
        ws.cell(row, 1, f"='設定'!B{EMP_FIRST_ROW + i}")
        ws.cell(row, 1).font = FONT_BOLD
        ws.cell(row, 1).border = THIN
        ws.cell(row, 1).fill = FILL_SETTINGS
        roster_row = name_row0 + i
        ws.cell(
            row,
            2,
            f'=IF(OR($E$5="",$E$5="非本年"),"",'
            f'IFERROR(INDEX(${first_col_letter}{roster_row}:${last_col_letter}{roster_row},$E$5),""))',
        )
        ws.cell(
            row,
            3,
            f'=IF(OR($E$6="",$E$6="非本年"),"",'
            f'IFERROR(INDEX(${first_col_letter}{roster_row}:${last_col_letter}{roster_row},$E$6),""))',
        )
        for col in (2, 3):
            ws.cell(row, col).border = THIN
            ws.cell(row, col).alignment = CENTER
            ws.cell(row, col).font = FONT_BOLD
            ws.cell(row, col).fill = FILL_QUICK

    # Legend matching main table colors
    ws["D8"] = "圖例"
    ws["D8"].font = FONT_BOLD
    for i, (code, fill) in enumerate(
        [("S", FILL_S), ("A", FILL_A), ("P", FILL_P), ("N", FILL_N), ("AL", FILL_AL), ("SL", FILL_SL)]
    ):
        cell = ws.cell(8, 5 + i, code)
        cell.fill = fill
        cell.border = THIN
        cell.alignment = CENTER
        cell.font = FONT_BOLD

    ws["A12"] = (
        "全年排班主表｜請假請到「請假登記」逐日登錄｜補假請到「假期上班登記」｜"
        "月曆檢視見「本月排班日曆」"
    )
    ws["A12"].font = FONT_BOLD
    ws.merge_cells("A12:F12")

    col_headers = [
        (1, "員工姓名"),
        (2, "已用年假\n(本年)"),
        (3, "剩餘年假\n(本年)"),
        (4, "累積補假\n結餘"),
        (5, "期初補假\n(可調)"),
        (6, "本年新增\n補假"),
    ]
    for col, text in col_headers:
        cell = ws.cell(header_row, col, text)
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN

    for w in range(NUM_WEEKS):
        col = FIRST_WEEK_COL + w
        cell = ws.cell(header_row, col)
        cell.value = week_start_formula(w)
        cell.number_format = "M/D"
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    for i, (name, role, offset, quota) in enumerate(EMPLOYEES):
        row = name_row0 + i
        set_row = EMP_FIRST_ROW + i

        ws.cell(row, 1, f"='設定'!B{set_row}")
        ws.cell(row, 1).font = FONT_BOLD
        ws.cell(row, 1).border = THIN
        ws.cell(row, 1).alignment = CENTER
        ws.cell(row, 1).fill = FILL_SETTINGS

        # AL used from 請假登記 by name + AL + year
        ws.cell(
            row,
            2,
            f'=IF(A{row}="","",'
            f'COUNTIFS(LeaveName,A{row},LeaveType,"AL",LeaveDate,">="&DATE({YEAR_CELL},1,1),'
            f'LeaveDate,"<="&DATE({YEAR_CELL},12,31)))',
        )
        # Remaining = personal quota - used
        ws.cell(
            row,
            3,
            f'=IF(A{row}="","",IF(\'設定\'!E{set_row}="","",\'設定\'!E{set_row}-B{row}))',
        )

        # Opening compensatory (manual)
        open_cell = ws.cell(row, 5, 0 if name else None)
        open_cell.fill = FILL_INPUT
        open_cell.border = THIN
        open_cell.alignment = CENTER

        # This year new comp from 假期上班登記
        ws.cell(row, 6, f'=IF(A{row}="","",CompYear{i + 1})')
        ws.cell(row, 6).border = THIN
        ws.cell(row, 6).fill = FILL_STAT
        ws.cell(row, 6).alignment = CENTER

        # Cumulative
        ws.cell(row, 4, f'=IF(A{row}="","",E{row}+F{row})')
        ws.cell(row, 4).border = THIN
        ws.cell(row, 4).fill = FILL_STAT
        ws.cell(row, 4).font = FONT_BOLD
        ws.cell(row, 4).alignment = CENTER

        for c in (2, 3):
            ws.cell(row, c).border = THIN
            ws.cell(row, c).fill = FILL_STAT
            ws.cell(row, c).alignment = CENTER

        # Week cells: leave in week overrides shift
        for w in range(NUM_WEEKS):
            col = FIRST_WEEK_COL + w
            col_l = get_column_letter(col)
            week_ref = f"{col_l}${header_row}"
            name_cell = f"'設定'!$B${set_row}"
            role_cell = f"'設定'!$C${set_row}"
            offset_cell = f"'設定'!$D${set_row}"
            # Any AL in week → AL; else any SL → SL; else planned
            has_al = (
                f'COUNTIFS(LeaveName,{name_cell},LeaveType,"AL",'
                f'LeaveDate,">="&{week_ref},LeaveDate,"<"&{week_ref}+7)'
            )
            has_sl = (
                f'COUNTIFS(LeaveName,{name_cell},LeaveType,"SL",'
                f'LeaveDate,">="&{week_ref},LeaveDate,"<"&{week_ref}+7)'
            )
            rot = rotate_shift_expr(week_ref, offset_cell)
            formula = (
                f'=IF({week_ref}="","",'
                f'IF({name_cell}="","",'
                f'IF({has_al}>0,"AL",'
                f'IF({has_sl}>0,"SL",'
                f'IF({role_cell}="S","S",'
                f'IF({role_cell}="R",{rot},""))))))'
            )
            cell = ws.cell(row, col, formula)
            cell.alignment = CENTER
            cell.border = THIN
            cell.font = FONT_BOLD

    # Holiday marker + helper count
    marker_row = name_row0 + 6
    hol_count_row = name_row0 + 7
    ws.cell(marker_row, 1, "公眾假期").font = FONT_SMALL
    ws.cell(marker_row, 1).fill = FILL_HOLIDAY
    ws.cell(marker_row, 1).border = THIN
    for w in range(NUM_WEEKS):
        col = FIRST_WEEK_COL + w
        col_l = get_column_letter(col)
        cell = ws.cell(marker_row, col)
        cell.value = (
            f'=IF({col_l}${header_row}="","",'
            f'IF(COUNTIFS(HolidayDates,">="&{col_l}${header_row},'
            f'HolidayDates,"<"&{col_l}${header_row}+7,HolidayDates,"<>")=0,"",'
            f'"假"&COUNTIFS(HolidayDates,">="&{col_l}${header_row},'
            f'HolidayDates,"<"&{col_l}${header_row}+7,HolidayDates,"<>")))'
        )
        cell.fill = FILL_HOLIDAY
        cell.font = FONT_SMALL
        cell.alignment = CENTER
        cell.border = THIN

        hcell = ws.cell(
            hol_count_row,
            col,
            f'=IF({col_l}${header_row}="","",'
            f'COUNTIFS(HolidayDates,">="&{col_l}${header_row},'
            f'HolidayDates,"<"&{col_l}${header_row}+7,HolidayDates,"<>"))',
        )
        hcell.font = FONT_SMALL
        hcell.fill = FILL_VACANT
    ws.cell(hol_count_row, 1, "（系統）").font = FONT_SMALL
    ws.row_dimensions[hol_count_row].hidden = True

    note_row = hol_count_row + 2
    ws.cell(note_row, 1, "使用提示（詳見「使用提示」工作表）").font = FONT_BOLD
    tips = [
        "1. 請假：只在「請假登記」新增列（姓名＋日期＋AL/SL）；主表與月曆會自動顯示。",
        "2. 補假：在「假期上班登記」對應假期列用下拉選擇上班人員；主表補假欄自動更新。",
        "3. 換年：改「設定」顯示年份，並在「公眾假期」追加假期；基準日 2026-06-29 勿改。",
        "4. 年假額度：在「設定」每人「年假額度（天）」各自填寫；剩餘年假＝個人額度−已用（按請假登記天數）。",
        "5. 同週若有超過 2 人出現 AL 或 SL（週欄顯示），整欄標紅警示。",
    ]
    for j, tip in enumerate(tips):
        ws.cell(note_row + 1 + j, 1, tip).font = FONT_SMALL
        ws.merge_cells(
            start_row=note_row + 1 + j,
            start_column=1,
            end_row=note_row + 1 + j,
            end_column=6,
        )

    for col, w in zip("ABCDEF", [12, 10, 10, 10, 10, 10]):
        ws.column_dimensions[col].width = w

    ws.row_dimensions[header_row].height = 32
    for r in range(name_row0, name_row0 + 6):
        ws.row_dimensions[r].height = 22

    ws.freeze_panes = "G14"

    # Conditional formatting — shifts + leave colors
    shift_range = f"{first_col_letter}{name_row0}:{last_col_letter}{name_row0 + 5}"
    add_shift_cf(ws, shift_range)
    add_shift_cf(ws, "B9:C14")

    for w in range(NUM_WEEKS):
        col_l = get_column_letter(FIRST_WEEK_COL + w)
        header_cell = f"{col_l}${header_row}"
        ws.conditional_formatting.add(
            f"{col_l}{header_row}",
            FormulaRule(
                formula=[
                    f'AND({header_cell}<>"",COUNTIFS(HolidayDates,">="&{header_cell},'
                    f'HolidayDates,"<"&{header_cell}+7,HolidayDates,"<>")>0)'
                ],
                fill=FILL_HOLIDAY,
            ),
        )
        col_range = f"{col_l}{header_row}:{col_l}{marker_row}"
        ws.conditional_formatting.add(
            col_range,
            FormulaRule(
                formula=[
                    f'(COUNTIF({col_l}${name_row0}:{col_l}${name_row0 + 5},"AL")+'
                    f'COUNTIF({col_l}${name_row0}:{col_l}${name_row0 + 5},"SL"))>2'
                ],
                fill=FILL_WARN,
            ),
        )

    # =====================================================================
    # 本月排班日曆（純檢視）
    # =====================================================================
    ws_cal = wb.create_sheet("本月排班日曆")
    ws_cal["A1"] = "本月排班日曆（純檢視｜請勿在此輸入）"
    ws_cal["A1"].font = FONT_TITLE
    ws_cal.merge_cells("A1:G1")

    ws_cal["A2"] = "目前月份"
    ws_cal["B2"] = '=TEXT(TODAY(),"YYYY年M月")'
    ws_cal["B2"].font = Font(name="Microsoft JhengHei", bold=True, size=14, color="1F4E79")
    ws_cal["C2"] = "← 隨 TODAY() 自動切換；請假異動請到「請假登記」"
    ws_cal["C2"].font = FONT_NOTE

    # Anchor: first day of current month, and Monday of that week
    ws_cal["A3"] = "本月1日"
    ws_cal["B3"] = '=DATE(YEAR(TODAY()),MONTH(TODAY()),1)'
    ws_cal["B3"].number_format = "YYYY-MM-DD"
    ws_cal["C3"] = "月曆起始週一"
    ws_cal["D3"] = "=B3-WEEKDAY(B3,2)+1"
    ws_cal["D3"].number_format = "YYYY-MM-DD"

    # Weekday headers
    cal_start_row = 5
    weekdays = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]
    for i, wd in enumerate(weekdays):
        cell = ws_cal.cell(cal_start_row, 1 + i, wd)
        cell.fill = FILL_HEADER
        cell.font = FONT_HDR
        cell.alignment = CENTER
        cell.border = THIN
        ws_cal.column_dimensions[get_column_letter(1 + i)].width = 16

    # Each day block: 8 rows (date, holiday name, 6 employees)
    # 6 week rows max
    BLOCK = 8
    for week in range(6):
        base_r = cal_start_row + 1 + week * BLOCK
        for dow in range(7):
            col = 1 + dow
            col_l = get_column_letter(col)
            # Date expression (no leading =) for embedding in other formulas
            d_expr = f"$D$3+{week}*7+{dow}"

            date_cell = ws_cal.cell(
                base_r,
                col,
                f'=IF(OR(MONTH({d_expr})<>MONTH($B$3),YEAR({d_expr})<>YEAR($B$3)),"",{d_expr})',
            )
            date_cell.number_format = "M/D"
            date_cell.font = FONT_WHITE
            date_cell.alignment = CENTER
            date_cell.border = THIN
            date_cell.fill = FILL_SUBHDR

            # Holiday name
            hol_cell = ws_cal.cell(
                base_r + 1,
                col,
                f'=IF({col_l}{base_r}="","",'
                f'IFERROR(INDEX(HolidayNames,MATCH({d_expr},HolidayDates,0)),""))',
            )
            hol_cell.font = FONT_TINY
            hol_cell.alignment = CENTER
            hol_cell.border = THIN
            hol_cell.fill = FILL_HOLIDAY

            # 6 employees — leave overrides planned shift
            for ei in range(6):
                er = base_r + 2 + ei
                set_row = EMP_FIRST_ROW + ei
                name_ref = f"'設定'!$B${set_row}"
                leave_al = (
                    f'COUNTIFS(LeaveName,{name_ref},LeaveDate,"="&({d_expr}),LeaveType,"AL")'
                )
                leave_sl = (
                    f'COUNTIFS(LeaveName,{name_ref},LeaveDate,"="&({d_expr}),LeaveType,"SL")'
                )
                cell = ws_cal.cell(
                    er,
                    col,
                    f'=IF(OR({col_l}{base_r}="",{name_ref}=""),"",'
                    f'IF({leave_al}>0,"AL",'
                    f'IF({leave_sl}>0,"SL",'
                    f'{planned_shift_formula(d_expr, set_row)})))',
                )
                cell.alignment = CENTER
                cell.border = THIN
                cell.font = FONT_SMALL
                cell.fill = FILL_READONLY

            # Holiday purple fill on date + holiday-name rows only (keep shift colors on emp rows)
            date_hol_range = f"{col_l}{base_r}:{col_l}{base_r + 1}"
            ws_cal.conditional_formatting.add(
                date_hol_range,
                FormulaRule(
                    formula=[f'AND({col_l}${base_r}<>"",{col_l}${base_r + 1}<>"")'],
                    fill=FILL_HOLIDAY,
                ),
            )
            emp_range = f"{col_l}{base_r + 2}:{col_l}{base_r + 7}"
            add_shift_cf(ws_cal, emp_range)

        # Employee name labels (right side) for this week block
        ws_cal.cell(base_r, 9, "日期").font = FONT_TINY
        ws_cal.cell(base_r + 1, 9, "公眾假期").font = FONT_TINY
        for ei in range(6):
            label = ws_cal.cell(
                base_r + 2 + ei,
                9,
                f"='設定'!B{EMP_FIRST_ROW + ei}",
            )
            label.font = FONT_TINY
            label.fill = FILL_SETTINGS
            label.border = THIN

    # Legend
    leg_r = cal_start_row + 1 + 6 * BLOCK + 1
    ws_cal.cell(leg_r, 1, "圖例（與排班表一致）").font = FONT_BOLD
    for i, (code, fill) in enumerate(
        [("S", FILL_S), ("A", FILL_A), ("P", FILL_P), ("N", FILL_N), ("AL", FILL_AL), ("SL", FILL_SL)]
    ):
        cell = ws_cal.cell(leg_r + 1, 1 + i, code)
        cell.fill = fill
        cell.border = THIN
        cell.alignment = CENTER
        cell.font = FONT_BOLD

    ws_cal.cell(leg_r + 3, 1, "注意：此表全部為公式自動生成，請勿直接輸入。請假請到「請假登記」。").font = FONT_NOTE
    ws_cal.merge_cells(start_row=leg_r + 3, start_column=1, end_row=leg_r + 3, end_column=7)
    ws_cal.column_dimensions["I"].width = 12
    ws_cal.freeze_panes = "A6"

    # =====================================================================
    # 使用提示
    # =====================================================================
    ws_help = wb.create_sheet("使用提示")
    ws_help["A1"] = "部門更表模版 — 使用提示"
    ws_help["A1"].font = FONT_TITLE
    ws_help.merge_cells("A1:B1")

    sections = [
        (
            "一、如何在「請假登記」新增請假紀錄",
            [
                "1. 打開「請假登記」工作表。",
                "2. 在表格空白列選擇【員工姓名】（下拉選單來自設定區）、填寫【請假日期】（實際請假當天）、選擇【假別】AL 或 SL。",
                "3. 連續多日請假：每一天各新增一列（例如 3/1～3/3 年假＝三列）。",
                "4. 排班表週欄、本月排班日曆、已用／剩餘年假都會自動用公式從本表計算，請勿在其他表輸入請假。",
                "5. 修改或刪除請假：直接改／清「請假登記」該列即可。",
            ],
        ),
        (
            "二、「本月排班日曆」為純檢視表（不可直接輸入）",
            [
                "1. 月曆隨 TODAY() 自動顯示當前月份（EOMONTH／DATE 推算）。",
                "2. 每日顯示六位員工當日班次（A/P/N/S）；若「請假登記」有該日紀錄則改顯示 AL／SL 並變色。",
                "3. 公眾假期會顯示假期名稱並以淺紫標示。",
                "4. 請勿在此表輸入任何資料；所有請假異動必須回到「請假登記」。",
                "5. 換月不會遺失紀錄——歷史請假永久保存在「請假登記」。",
            ],
        ),
        (
            "三、如何在「假期上班登記」手動登記上班人員",
            [
                "1. 打開「假期上班登記」：日期與假期名稱已由「公眾假期」自動帶入。",
                "2. 在該列「上班人員1～6」用下拉選單選擇當日實際仍上班的員工（可留空）。",
                "3. 下方「補假統計區」會自動 COUNTIF／SUMPRODUCT 計算每人公眾假期上班次數。",
                "4. 「排班表」的本年新增補假／累積補假結餘會引用此統計（累積＝期初補假＋本年）。",
                "5. 換年前建議把累積結餘抄入「期初補假」，再改顯示年份。",
            ],
        ),
        (
            "四、年頭更新與人員",
            [
                "1. 「設定」→ 只改【顯示年份】；【輪班基準日】保持 2026-06-29。",
                "2. 「公眾假期」追加新一年日期與名稱。",
                "3. 每人【年假額度（天）】在設定表各自填寫（已取消統一額度）。",
                "4. 改姓名只改設定區；全表公式自動更新。第六人：填姓名、類型改 R/S、設 Offset 與額度。",
            ],
        ),
        (
            "五、16 週輪班（不變）",
            [
                "錨點 2026-06-29；MOD(週差+Offset,16)：0–3 A、4–7 P、8–11 N、12–15 P。",
                "Offset 0/4/8/12 錯開，保證每週 A/P/N 皆有人。ARTHUR 固定 S。",
            ],
        ),
        (
            "六、顏色圖例",
            [
                "S＝白｜A＝灰｜P＝綠｜N＝藍｜AL＝黃｜SL＝橙｜公眾假期＝淺紫｜衝突週＝紅",
            ],
        ),
    ]

    r = 3
    for title, lines in sections:
        ws_help.cell(r, 1, title).fill = FILL_SUBHDR
        ws_help.cell(r, 1).font = FONT_WHITE
        ws_help.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r += 1
        for line in lines:
            ws_help.cell(r, 1, line).font = FONT_NORMAL
            ws_help.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            r += 1
        r += 1

    ws_help.column_dimensions["A"].width = 100

    # Sheet order
    order = [
        "排班表",
        "設定",
        "請假登記",
        "本月排班日曆",
        "假期上班登記",
        "公眾假期",
        "使用提示",
    ]
    for idx, name in enumerate(order):
        wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    out = "/workspace/部門更表模版_跨年.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    path = build_workbook()
    print(f"Created: {path}")
