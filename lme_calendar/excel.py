"""Write/update only the three LME calendar sheets inside a workbook."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from lme_calendar.dates import (
    calc_3m_date,
    calc_cash_date,
    generate_trading_calendar,
    is_business_day,
    iter_calendar_days,
    weekday_name,
)
from lme_calendar.holidays import HolidayCalendar

MANAGED_SHEETS = ("Trading_Calendar", "Cash_3M_Daily", "Holiday_Reference")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
TRUE_FILL = PatternFill("solid", fgColor="C6EFCE")
FALSE_FILL = PatternFill("solid", fgColor="FFC7CE")
ROW_FILL_ALT = PatternFill("solid", fgColor="D6EAF8")
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

TRADING_HEADERS = (
    "Date",
    "Weekday",
    "Is_UK_Holiday",
    "Is_US_Holiday",
    "Is_Valid_Trading_Day",
)
CASH_HEADERS = ("Snapshot_Date", "Cash_Date", "3M_Date", "Days_Cash_to_3M")
HOLIDAY_HEADERS = ("Date", "Region", "Holiday_Name")


def _style_header(ws: Worksheet, headers: tuple[str, ...]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(1, col, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN
    ws.row_dimensions[1].height = 22


def _clear_data_rows(ws: Worksheet) -> None:
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def _write_bool(ws: Worksheet, row: int, col: int, value: bool) -> None:
    cell = ws.cell(row, col, value)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = TRUE_FILL if value else FALSE_FILL
    cell.border = THIN


def _write_date(ws: Worksheet, row: int, col: int, value: date) -> None:
    cell = ws.cell(row, col, value)
    cell.number_format = "YYYY-MM-DD"
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN


def _write_text(ws: Worksheet, row: int, col: int, value: object) -> None:
    cell = ws.cell(row, col, value)
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN


def _autosize(ws: Worksheet, min_width: int = 14) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[letter].width = max(min_width, length + 4)


def _sheet(wb: Workbook, name: str) -> Worksheet:
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def _set_named_range(wb: Workbook, name: str, sheet: str, cell: str) -> None:
    defn = DefinedName(name, attr_text=f"'{sheet}'!{cell}")
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(defn)


def write_holiday_reference(ws: Worksheet, calendar: HolidayCalendar) -> None:
    _clear_data_rows(ws)
    _style_header(ws, HOLIDAY_HEADERS)
    for row, item in enumerate(calendar.holidays, start=2):
        _write_date(ws, row, 1, item.date)
        _write_text(ws, row, 2, item.region)
        _write_text(ws, row, 3, item.name)
    _autosize(ws)


def write_trading_calendar(
    ws: Worksheet,
    *,
    start_date: date,
    months_forward: int,
    calendar: HolidayCalendar,
    retain_expired: bool,
) -> int:
    """Replace or merge the rolling window. Returns the number of valid trading days."""
    window_days = iter_calendar_days(start_date, months_forward)
    window_start, window_end = window_days[0], window_days[-1]

    retained: list[tuple] = []
    if retain_expired and ws.max_row >= 2:
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row, 1).value
            if isinstance(value, datetime):
                value = value.date()
            if isinstance(value, date) and value < window_start:
                retained.append(
                    (
                        value,
                        ws.cell(row, 2).value,
                        bool(ws.cell(row, 3).value),
                        bool(ws.cell(row, 4).value),
                        bool(ws.cell(row, 5).value),
                    )
                )

    rows: list[tuple] = list(retained)
    for day in window_days:
        uk = calendar.is_uk_holiday(day)
        us = calendar.is_us_holiday(day)
        valid = is_business_day(day, calendar)
        rows.append((day, weekday_name(day), uk, us, valid))

    _clear_data_rows(ws)
    _style_header(ws, TRADING_HEADERS)
    for excel_row, (day, weekday, uk, us, valid) in enumerate(rows, start=2):
        _write_date(ws, excel_row, 1, day)
        _write_text(ws, excel_row, 2, weekday)
        _write_bool(ws, excel_row, 3, uk)
        _write_bool(ws, excel_row, 4, us)
        _write_bool(ws, excel_row, 5, valid)
        if not valid:
            ws.cell(excel_row, 2).fill = FALSE_FILL
    _autosize(ws)
    ws.auto_filter.ref = f"A1:E{max(1, ws.max_row)}"
    return sum(1 for day in generate_trading_calendar(start_date, months_forward, calendar))


def write_cash_3m(
    ws: Worksheet,
    *,
    snapshot_date: date,
    cash_date: date,
    three_m_date: date,
    snapshot_mode: str,
) -> int:
    if snapshot_mode not in {"overwrite", "append"}:
        raise ValueError("snapshot_mode must be 'overwrite' or 'append'")

    if ws.max_row < 1 or ws.cell(1, 1).value != CASH_HEADERS[0]:
        _clear_data_rows(ws)
        _style_header(ws, CASH_HEADERS)

    days_between = (three_m_date - cash_date).days
    record = (snapshot_date, cash_date, three_m_date, days_between)

    if snapshot_mode == "overwrite":
        _clear_data_rows(ws)
        _style_header(ws, CASH_HEADERS)
        target_row = 2
    else:
        if ws.max_row == 1:
            target_row = 2
        else:
            last_snapshot = ws.cell(ws.max_row, 1).value
            if isinstance(last_snapshot, datetime):
                last_snapshot = last_snapshot.date()
            if last_snapshot == snapshot_date:
                target_row = ws.max_row  # same-day rerun replaces that row
            else:
                target_row = ws.max_row + 1

    _write_date(ws, target_row, 1, record[0])
    _write_date(ws, target_row, 2, record[1])
    _write_date(ws, target_row, 3, record[2])
    _write_text(ws, target_row, 4, record[3])
    for col in range(1, 5):
        ws.cell(target_row, col).fill = ROW_FILL_ALT
        ws.cell(target_row, col).border = THIN
    _autosize(ws)
    ws.auto_filter.ref = f"A1:D{max(1, ws.max_row)}"
    return target_row


def update_workbook(
    excel_path: str | Path,
    *,
    calendar: HolidayCalendar,
    as_of: date,
    months_forward: int = 27,
    snapshot_mode: str = "overwrite",
    retain_expired: bool = False,
) -> dict:
    path = Path(excel_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.is_file():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        default = wb.active
        default.title = "Trading_Calendar"

    # Never delete unmanaged sheets (Bloomberg BDP/BDH tabs, etc.).
    for name in MANAGED_SHEETS:
        _sheet(wb, name)

    # Drop the leftover default sheet only when it is unused and unmanaged.
    if (
        "Sheet" in wb.sheetnames
        and "Sheet" not in MANAGED_SHEETS
        and wb["Sheet"].max_row == 1
        and wb["Sheet"].max_column == 1
        and wb["Sheet"].cell(1, 1).value is None
    ):
        wb.remove(wb["Sheet"])

    cash_date = calc_cash_date(as_of, calendar)
    three_m_date = calc_3m_date(as_of, calendar)

    write_holiday_reference(wb["Holiday_Reference"], calendar)
    valid_days = write_trading_calendar(
        wb["Trading_Calendar"],
        start_date=as_of,
        months_forward=months_forward,
        calendar=calendar,
        retain_expired=retain_expired,
    )
    snapshot_row = write_cash_3m(
        wb["Cash_3M_Daily"],
        snapshot_date=as_of,
        cash_date=cash_date,
        three_m_date=three_m_date,
        snapshot_mode=snapshot_mode,
    )

    _set_named_range(wb, "Snapshot_Date", "Cash_3M_Daily", f"$A${snapshot_row}")
    _set_named_range(wb, "Cash_Date", "Cash_3M_Daily", f"$B${snapshot_row}")
    _set_named_range(wb, "ThreeM_Date", "Cash_3M_Daily", f"$C${snapshot_row}")
    _set_named_range(wb, "Days_Cash_to_3M", "Cash_3M_Daily", f"$D${snapshot_row}")

    # Keep a stable sheet order without moving unmanaged sheets around more than needed.
    desired = [name for name in MANAGED_SHEETS if name in wb.sheetnames]
    for index, name in enumerate(desired):
        current = wb.sheetnames.index(name)
        if current != index:
            wb.move_sheet(name, offset=index - current)

    wb.save(path)
    return {
        "excel_path": str(path),
        "snapshot_date": as_of,
        "cash_date": cash_date,
        "three_m_date": three_m_date,
        "days_cash_to_3m": (three_m_date - cash_date).days,
        "valid_trading_days": valid_days,
        "snapshot_mode": snapshot_mode,
        "snapshot_row": snapshot_row,
        "holiday_count": len(calendar.holidays),
        "coverage_start": calendar.coverage_start,
        "coverage_end": calendar.coverage_end,
    }
